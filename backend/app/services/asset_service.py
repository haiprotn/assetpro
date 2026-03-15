"""
Asset Service - Business Logic Layer
"""
import uuid
import json
import secrets
from datetime import datetime, date, timedelta
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, text
from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from app.models.assets import Asset, AssetTypeGroup, AttributeDefinition, AssetLifecycleEvent
from app.schemas.assets import AssetCreate, AssetUpdate, AssetListOut, AssetOut


class AssetService:

    @staticmethod
    async def create_asset(db: AsyncSession, data: AssetCreate, created_by: uuid.UUID) -> Asset:
        # Validate dynamic attributes against group schema
        await AssetService._validate_dynamic_attributes(db, data.asset_type_id, data.dynamic_attributes)

        # Generate QR code
        qr_token = f"ASSET-{data.asset_code}-{secrets.token_urlsafe(8).upper()}"

        asset = Asset(
            id=uuid.uuid4(),
            qr_code=qr_token,
            created_by=created_by,
            **data.model_dump()
        )
        db.add(asset)
        await db.commit()
        await db.refresh(asset)
        return asset

    @staticmethod
    async def _validate_dynamic_attributes(
        db: AsyncSession, asset_type_id: uuid.UUID, attributes: dict
    ):
        """
        Validate dynamic_attributes against AttributeDefinition rules for the group.
        Raises HTTPException with field-level errors if invalid.
        """
        # Get required fields for this asset type's group
        stmt = text("""
            SELECT ad.field_key, ad.field_label, ad.field_type, 
                   ad.is_required, ad.validation_rules, ad.select_options
            FROM attribute_definitions ad
            JOIN asset_types at ON at.group_id = ad.group_id
            WHERE at.id = :type_id
        """)
        result = await db.execute(stmt, {"type_id": str(asset_type_id)})
        attr_defs = result.fetchall()

        errors = []
        for attr in attr_defs:
            val = attributes.get(attr.field_key)

            # Required check
            if attr.is_required and val is None:
                errors.append(f"'{attr.field_label}' là bắt buộc")
                continue

            if val is None:
                continue

            # Type validation
            if attr.field_type == 'NUMBER':
                try:
                    num_val = float(val)
                    rules = attr.validation_rules or {}
                    if 'min' in rules and num_val < rules['min']:
                        errors.append(f"'{attr.field_label}' phải >= {rules['min']}")
                    if 'max' in rules and num_val > rules['max']:
                        errors.append(f"'{attr.field_label}' phải <= {rules['max']}")
                except (TypeError, ValueError):
                    errors.append(f"'{attr.field_label}' phải là số")

            elif attr.field_type == 'DATE':
                try:
                    datetime.strptime(str(val), '%Y-%m-%d')
                except ValueError:
                    errors.append(f"'{attr.field_label}' phải có định dạng YYYY-MM-DD")

            elif attr.field_type == 'SELECT':
                options = attr.select_options or []
                if val not in options:
                    errors.append(f"'{attr.field_label}' phải là một trong: {', '.join(options)}")

        if errors:
            raise HTTPException(status_code=422, detail={"dynamic_attribute_errors": errors})

    @staticmethod
    async def list_assets(
        db: AsyncSession,
        q: Optional[str] = None,
        status: Optional[str] = None,
        asset_type_id: Optional[uuid.UUID] = None,
        group_code: Optional[str] = None,
        location_id: Optional[uuid.UUID] = None,
        department_id: Optional[uuid.UUID] = None,
        page: int = 1,
        size: int = 20,
        sort_by: Optional[str] = None,
        sort_dir: Optional[str] = "asc",
    ) -> AssetListOut:
        where_parts = ["a.is_active = TRUE"]
        params: dict = {}

        if q:
            where_parts.append(
                "(a.name ILIKE :q OR a.asset_code ILIKE :q OR a.barcode ILIKE :q OR a.chassis_number ILIKE :q OR s.name ILIKE :q)"
            )
            params["q"] = f"%{q}%"
        if status:
            where_parts.append("a.status = :status")
            params["status"] = status
        if asset_type_id:
            where_parts.append("a.asset_type_id = :asset_type_id")
            params["asset_type_id"] = str(asset_type_id)
        if group_code:
            where_parts.append("atg.code = :group_code")
            params["group_code"] = group_code
        if location_id:
            where_parts.append("a.current_location_id = :location_id")
            params["location_id"] = str(location_id)
        if department_id:
            where_parts.append("a.managing_department_id = :department_id")
            params["department_id"] = str(department_id)

        where = " AND ".join(where_parts)
        joins = (
            "LEFT JOIN asset_types    at2 ON at2.id  = a.asset_type_id"
            " LEFT JOIN asset_type_groups atg ON atg.id  = at2.group_id"
            " LEFT JOIN locations       l   ON l.id   = a.current_location_id"
            " LEFT JOIN departments     d   ON d.id   = a.managing_department_id"
            " LEFT JOIN suppliers       s   ON s.id   = a.supplier_id"
        )

        total = (await db.execute(
            text(f"SELECT COUNT(*) FROM assets a {joins} WHERE {where}"),
            params,
        )).scalar()

        params["limit_val"] = size
        params["offset_val"] = (page - 1) * size

        _SORT_MAP = {
            "name": "a.name", "status": "a.status",
            "location": "l.name", "department": "d.name",
            "original_value": "a.original_value", "purchase_date": "a.purchase_date",
            "asset_type": "at2.name",
        }
        order_dir = "ASC" if str(sort_dir).lower() == "asc" else "DESC"
        if sort_by is None:
            order_dir = "DESC"  # default newest first

        if sort_by == "asset_code":
            order_clause = (
                f"regexp_replace(a.asset_code, '[0-9]+', '', 'g') {order_dir}, "
                f"NULLIF(regexp_replace(a.asset_code, '[^0-9]+', '', 'g'), '')::BIGINT {order_dir} NULLS LAST"
            )
        else:
            order_col = _SORT_MAP.get(sort_by, "a.created_at")
            order_clause = f"{order_col} {order_dir} NULLS LAST"

        rows = (await db.execute(text(f"""
            SELECT
                a.*,
                at2.name AS asset_type_name,
                l.name   AS location_name,
                d.name   AS department_name,
                s.name   AS supplier_name
            FROM assets a
            {joins}
            WHERE {where}
            ORDER BY {order_clause}
            LIMIT :limit_val OFFSET :offset_val
        """), params)).mappings().all()

        items = [AssetOut.model_validate(dict(r)) for r in rows]
        return AssetListOut(total=total, page=page, size=size, items=items)

    _DETAIL_JOINS = (
        " LEFT JOIN asset_types  at2 ON at2.id = a.asset_type_id"
        " LEFT JOIN locations    l   ON l.id   = a.current_location_id"
        " LEFT JOIN departments  d   ON d.id   = a.managing_department_id"
        " LEFT JOIN suppliers    s   ON s.id   = a.supplier_id"
    )
    _DETAIL_SELECT = (
        "SELECT a.*, at2.name AS asset_type_name, l.name AS location_name,"
        " d.name AS department_name, s.name AS supplier_name FROM assets a"
    )

    @staticmethod
    async def get_by_id(db: AsyncSession, asset_id: uuid.UUID):
        row = (await db.execute(
            text(f"{AssetService._DETAIL_SELECT}{AssetService._DETAIL_JOINS} WHERE a.id = :id"),
            {"id": str(asset_id)},
        )).mappings().one_or_none()
        return AssetOut.model_validate(dict(row)) if row else None

    @staticmethod
    async def get_by_code(db: AsyncSession, code: str):
        row = (await db.execute(
            text(f"{AssetService._DETAIL_SELECT}{AssetService._DETAIL_JOINS} WHERE a.asset_code = :code"),
            {"code": code},
        )).mappings().one_or_none()
        return AssetOut.model_validate(dict(row)) if row else None

    @staticmethod
    async def update_asset(
        db: AsyncSession, asset_id: uuid.UUID, data: AssetUpdate, updated_by: uuid.UUID
    ):
        check = await db.execute(select(Asset).where(Asset.id == asset_id))
        orm_asset = check.scalar_one_or_none()
        if not orm_asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(orm_asset, field, value)
        orm_asset.updated_by = updated_by

        await db.commit()
        return await AssetService.get_by_id(db, asset_id)

    @staticmethod
    async def delete_asset(db: AsyncSession, asset_id: uuid.UUID):
        check = await db.execute(select(Asset).where(Asset.id == asset_id))
        orm_asset = check.scalar_one_or_none()
        if not orm_asset:
            raise HTTPException(status_code=404, detail="Asset not found")
        orm_asset.is_active = False
        await db.commit()

    @staticmethod
    async def duplicate_asset(db: AsyncSession, asset_id: uuid.UUID, created_by: uuid.UUID):
        orig = (await db.execute(select(Asset).where(Asset.id == asset_id))).scalar_one_or_none()
        if not orig:
            raise HTTPException(status_code=404, detail="Asset not found")
        base_code = orig.asset_code
        new_code = f"{base_code}_COPY"
        for i in range(1, 100):
            candidate = new_code if i == 1 else f"{base_code}_COPY{i}"
            exists = (await db.execute(select(Asset).where(Asset.asset_code == candidate))).scalar_one_or_none()
            if not exists:
                new_code = candidate
                break
        new_asset = Asset(
            id=uuid.uuid4(),
            asset_code=new_code,
            name=f"[Nhân bản] {orig.name}",
            status='PENDING_ALLOCATION',
            asset_type_id=orig.asset_type_id,
            legal_entity_id=orig.legal_entity_id,
            managing_department_id=orig.managing_department_id,
            current_location_id=orig.current_location_id,
            supplier_id=orig.supplier_id,
            model_series=orig.model_series,
            year_manufactured=orig.year_manufactured,
            country_manufactured=orig.country_manufactured,
            purchase_price=orig.purchase_price,
            original_value=orig.original_value,
            depreciation_months=orig.depreciation_months,
            loan_amount=orig.loan_amount,
            purchase_date=orig.purchase_date,
            warranty_end_date=orig.warranty_end_date,
            expiry_date=orig.expiry_date,
            warranty_months=orig.warranty_months,
            registration_expiry=orig.registration_expiry,
            quantity=orig.quantity,
            description=orig.description,
            condition_description=orig.condition_description,
            tags=orig.tags,
            dynamic_attributes=orig.dynamic_attributes or {},
            created_by=created_by,
        )
        db.add(new_asset)
        await db.commit()
        return await AssetService.get_by_id(db, new_asset.id)

    @staticmethod
    async def upload_image(db: AsyncSession, asset_id: uuid.UUID, file, upload_dir: str):
        import os
        check = await db.execute(select(Asset).where(Asset.id == asset_id))
        orm_asset = check.scalar_one_or_none()
        if not orm_asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
        filename = f"{orm_asset.asset_code}{ext}"
        filepath = os.path.join(upload_dir, filename)
        os.makedirs(upload_dir, exist_ok=True)
        content = await file.read()
        with open(filepath, "wb") as f:
            f.write(content)

        orm_asset.asset_image_url = f"/uploads/assets/{filename}"
        await db.commit()
        return await AssetService.get_by_id(db, asset_id)

    @staticmethod
    async def get_alerts(db: AsyncSession):
        today = date.today()
        in_30_days = today + timedelta(days=30)

        # Registration expiry alerts
        reg_stmt = select(Asset).where(and_(
            Asset.registration_expiry.isnot(None),
            Asset.registration_expiry <= in_30_days,
            Asset.status == 'IN_USE'
        ))
        reg_expiring = (await db.execute(reg_stmt)).scalars().all()

        # Calibration expiry (dynamic attribute)
        calib_stmt = text("""
            SELECT id, asset_code, name, dynamic_attributes->>'han_hieu_chuan' as calib_date
            FROM assets
            WHERE dynamic_attributes->>'han_hieu_chuan' IS NOT NULL
            AND (dynamic_attributes->>'han_hieu_chuan')::date <= :cutoff
            AND is_active = TRUE
        """)
        calib_result = await db.execute(calib_stmt, {"cutoff": in_30_days})
        calib_expiring = calib_result.fetchall()

        return {
            "registration_expiry": [
                {"id": str(a.id), "code": a.asset_code, "name": a.name, "expiry": a.registration_expiry}
                for a in reg_expiring
            ],
            "calibration_expiry": [
                {"id": str(r.id), "code": r.asset_code, "name": r.name, "expiry": r.calib_date}
                for r in calib_expiring
            ],
        }

    @staticmethod
    async def get_stats_by_location(db: AsyncSession):
        stmt = text("""
            SELECT 
                l.name as location_name,
                COUNT(a.id) as asset_count,
                SUM(a.original_value) as total_value,
                SUM(CASE WHEN a.status = 'IN_USE' THEN 1 ELSE 0 END) as in_use_count
            FROM assets a
            LEFT JOIN locations l ON a.current_location_id = l.id
            WHERE a.is_active = TRUE
            GROUP BY l.name
            ORDER BY asset_count DESC
        """)
        result = await db.execute(stmt)
        rows = result.fetchall()
        return [dict(r._mapping) for r in rows]

    @staticmethod
    async def get_lifecycle(db: AsyncSession, asset_id: uuid.UUID):
        stmt = (
            select(AssetLifecycleEvent)
            .where(AssetLifecycleEvent.asset_id == asset_id)
            .order_by(AssetLifecycleEvent.created_at.desc())
        )
        events = (await db.execute(stmt)).scalars().all()
        return events

    @staticmethod
    async def generate_qr_image(db: AsyncSession, asset_id: uuid.UUID):
        import qrcode
        import io

        asset = await AssetService.get_by_id(db, asset_id)
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr_data = json.dumps({
            "type": "ASSET",
            "code": asset.asset_code,
            "qr": asset.qr_code,
        })
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")

    @staticmethod
    async def add_attachment(db, asset_id, file, attachment_type, uploaded_by):
        # Implementation: upload to S3/MinIO, store URL in DB
        raise NotImplementedError("File upload service not configured")

    @staticmethod
    async def get_attachments(db, asset_id):
        from app.models.assets import AssetAttachment
        stmt = select(AssetAttachment).where(AssetAttachment.asset_id == asset_id)
        items = (await db.execute(stmt)).scalars().all()
        return items

    @staticmethod
    async def export_assets(
        db: AsyncSession,
        format: str = "xlsx",
        columns: Optional[List[str]] = None,
        status: Optional[str] = None,
        group_code: Optional[str] = None,
    ):
        import io
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.drawing.image import Image as XlImage
        _upload_dir = os.environ.get("UPLOAD_DIR", "/app/uploads/assets")

        where_clauses = ["a.is_active = TRUE"]
        params: dict = {}
        if status:
            where_clauses.append("a.status = :status")
            params["status"] = status
        if group_code:
            where_clauses.append("atg.code = :group_code")
            params["group_code"] = group_code
        where_sql = " AND ".join(where_clauses)

        stmt = text(f"""
            SELECT
                a.asset_code, a.barcode, a.qr_code, a.name, a.status,
                a.quantity, a.qty_maintenance, a.qty_liquidated, a.qty_lost,
                a.qty_cancelled, a.qty_broken, a.qty_allocated,
                a.model_series, a.year_manufactured, a.country_manufactured,
                a.purchase_price, a.original_value, a.current_value,
                a.depreciation_value, a.depreciation_months, a.loan_amount,
                a.purchase_date, a.report_increase_date, a.warranty_end_date,
                a.expiry_date, a.warranty_months, a.chassis_number, a.engine_number,
                a.license_plate, a.registration_expiry, a.condition_description, a.description,
                a.dynamic_attributes, a.asset_image_url, a.attachment_count, a.created_at,
                le.name   AS legal_entity_name,
                at2.name  AS asset_type_name,
                atg.name  AS asset_type_group_name,
                l.name    AS location_name,
                d.name    AS department_name,
                s.name    AS supplier_name,
                s.address AS supplier_address,
                s.phone   AS supplier_phone
            FROM assets a
            LEFT JOIN legal_entities le     ON a.legal_entity_id       = le.id
            LEFT JOIN asset_types at2       ON a.asset_type_id          = at2.id
            LEFT JOIN asset_type_groups atg ON at2.group_id             = atg.id
            LEFT JOIN locations l           ON a.current_location_id    = l.id
            LEFT JOIN departments d         ON a.managing_department_id = d.id
            LEFT JOIN suppliers s           ON a.supplier_id            = s.id
            WHERE {where_sql}
            ORDER BY a.created_at DESC
        """)

        result = await db.execute(stmt, params)
        rows = result.fetchall()

        STATUS_MAP = {
            'PENDING_ALLOCATION': 'Chờ cấp phát',
            'IN_USE': 'Đang sử dụng',
            'IN_MAINTENANCE': 'Đang bảo trì',
            'LIQUIDATED': 'Thanh lý',
            'LOST': 'Mất',
            'BROKEN': 'Hỏng',
            'CANCELLED': 'Huỷ',
        }

        def img_name(url):
            return os.path.basename(url) if url else ''

        def sf(v):
            try:
                return float(v) if v is not None else ''
            except Exception:
                return ''

        def dyn(r, key):
            try:
                return (r.dynamic_attributes or {}).get(key, '') or ''
            except Exception:
                return ''

        ALL_COL_DEFS = [
            ('phap_nhan',            'Pháp nhân sở hữu',      lambda r: r.legal_entity_name or ''),
            ('ma_qrcode',            'Mã QRCode',              lambda r: r.qr_code or ''),
            ('anh_tai_san',          'Ảnh tài sản',            lambda r: img_name(r.asset_image_url)),
            ('ma_ts',                'Mã TS',                  lambda r: r.asset_code),
            ('barcode',              'Barcode',                lambda r: r.barcode or ''),
            ('ten_ts',               'Tên tài sản',            lambda r: r.name),
            ('loai_ts',              'Loại tài sản',           lambda r: r.asset_type_name or ''),
            ('nhom_ts',              'Nhóm tài sản',           lambda r: r.asset_type_group_name or ''),
            ('trang_thai',           'Trạng thái',             lambda r: STATUS_MAP.get(r.status, r.status)),
            ('vi_tri',               'Vị trí tài sản',         lambda r: r.location_name or ''),
            ('phong_ban',            'Phòng ban quản lý',      lambda r: r.department_name or ''),
            ('model_series',         'Model / Series',         lambda r: r.model_series or ''),
            ('mo_ta',                'Mô tả',                  lambda r: r.description or ''),
            ('dien_giai_tinh_trang', 'Diễn giải tình trạng',  lambda r: r.condition_description or ''),
            ('ngay_tao',             'Ngày tạo',               lambda r: str(r.created_at.date()) if r.created_at else ''),
            ('so_luong',             'Số lượng',               lambda r: r.quantity),
            ('sl_bao_tri',           'SL Bảo trì',             lambda r: r.qty_maintenance),
            ('sl_con_lai',           'SL còn lại',             lambda r: r.quantity - r.qty_allocated - r.qty_liquidated),
            ('sl_thanh_ly',          'SL thanh lý',            lambda r: r.qty_liquidated),
            ('sl_mat',               'SL mất',                 lambda r: r.qty_lost),
            ('sl_huy',               'SL huỷ',                 lambda r: r.qty_cancelled),
            ('sl_hong',              'SL hỏng',                lambda r: r.qty_broken),
            ('nguyen_gia',           'Nguyên giá',             lambda r: sf(r.original_value)),
            ('gia_mua',              'Giá mua',                lambda r: sf(r.purchase_price)),
            ('gia_tri_hien_tai',     'Giá trị hiện tại',       lambda r: sf(r.current_value)),
            ('gia_tri_khau_hao',     'Giá trị khấu hao',       lambda r: sf(r.depreciation_value)),
            ('khau_hao_thang',       'Khấu hao (tháng)',       lambda r: r.depreciation_months),
            ('da_vay',               'Đã vay',                 lambda r: sf(r.loan_amount)),
            ('ngay_mua',             'Ngày mua',               lambda r: str(r.purchase_date) if r.purchase_date else ''),
            ('ngay_bao_tang',        'Ngày báo tăng',          lambda r: str(r.report_increase_date) if r.report_increase_date else ''),
            ('ngay_ket_thuc_bh',     'Ngày kết thúc BH',       lambda r: str(r.warranty_end_date) if r.warranty_end_date else ''),
            ('ngay_het_han',         'Ngày hết hạn',           lambda r: str(r.expiry_date) if r.expiry_date else ''),
            ('thoi_gian_bh',         'Thời gian BH (tháng)',   lambda r: r.warranty_months),
            ('han_dang_kiem',        'Hạn đăng kiểm',          lambda r: str(r.registration_expiry) if r.registration_expiry else ''),
            ('han_hieu_chuan',       'Hạn hiệu chuẩn',         lambda r: dyn(r, 'han_hieu_chuan')),
            ('nha_cung_cap',         'Nhà cung cấp',           lambda r: r.supplier_name or ''),
            ('dia_chi_ncc',          'Địa chỉ NCC',            lambda r: r.supplier_address or ''),
            ('dien_thoai_ncc',       'Điện thoại NCC',         lambda r: r.supplier_phone or ''),
            ('so_khung',             'Số Khung',               lambda r: r.chassis_number or ''),
            ('so_dong_co',           'Số Động cơ',             lambda r: r.engine_number or ''),
            ('bien_so',              'Biển số xe',             lambda r: r.license_plate or ''),
            ('nam_nuoc_sx',          'Năm - Nước sản xuất',    lambda r: f"{r.year_manufactured or ''} {r.country_manufactured or ''}".strip()),
            ('thue_bao',             'Thuê bao',               lambda r: dyn(r, 'thue_bao')),
            ('thong_so_thiet_bi',    'Thông số thiết bị',      lambda r: dyn(r, 'thong_so_thiet_bi')),
            ('so_file_dinh_kem',     'Số file đính kèm',       lambda r: r.attachment_count),
        ]

        if columns:
            key_order = {k: i for i, k in enumerate(columns)}
            col_defs = sorted(
                [c for c in ALL_COL_DEFS if c[0] in key_order],
                key=lambda c: key_order[c[0]]
            )
        else:
            col_defs = ALL_COL_DEFS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tài sản"
        ws.row_dimensions[1].height = 28

        header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for ci, (_, label, _) in enumerate(col_defs, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        IMG_ROW_H = 55  # px row height when image present

        for ri, row in enumerate(rows, 2):
            has_img = False
            for ci, (col_key, _, extractor) in enumerate(col_defs, 1):
                if col_key == 'anh_tai_san' and row.asset_image_url:
                    img_path = os.path.join(_upload_dir, os.path.basename(row.asset_image_url))
                    if os.path.exists(img_path):
                        try:
                            xl_img = XlImage(img_path)
                            xl_img.width, xl_img.height = 60, 50
                            cell_addr = ws.cell(row=ri, column=ci).coordinate
                            ws.add_image(xl_img, cell_addr)
                            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 10
                            has_img = True
                        except Exception:
                            ws.cell(row=ri, column=ci, value=os.path.basename(row.asset_image_url))
                    else:
                        ws.cell(row=ri, column=ci, value=os.path.basename(row.asset_image_url))
                else:
                    try:
                        val = extractor(row)
                    except Exception:
                        val = ''
                    ws.cell(row=ri, column=ci, value=val)
            if has_img:
                ws.row_dimensions[ri].height = IMG_ROW_H

        for ci, (col_key, _, _) in enumerate(col_defs, 1):
            if col_key == 'anh_tai_san':
                continue
            col_letter = ws.cell(row=1, column=ci).column_letter
            max_len = max((len(str(ws.cell(row=r, column=ci).value or '')) for r in range(1, len(rows) + 2)), default=8)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"tai-san-{date.today().strftime('%d%m%Y')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
