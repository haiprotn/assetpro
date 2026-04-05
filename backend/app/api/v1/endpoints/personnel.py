"""
Personnel API Endpoints - Quản lý nhân sự
"""
import io
import math
import os
import uuid
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload

from app.db.session import get_db
from app.core.auth import get_current_user
from app.models.assets import Personnel, Department
from app.models.personnel import Position, PositionTitle, ContractType, EmployeeContract, PersonnelDocument
from app.schemas.personnel import (
    PersonnelCreate, PersonnelUpdate, PersonnelOut, PersonnelDetail, PaginatedPersonnel,
    PositionCreate, PositionUpdate, PositionOut,
    PositionTitleCreate, PositionTitleUpdate, PositionTitleOut,
    ContractTypeCreate, ContractTypeUpdate, ContractTypeOut,
    EmployeeContractCreate, EmployeeContractUpdate, EmployeeContractOut,
)

_PERSONNEL_UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "/app/uploads/assets").replace("/assets", "/personnel")

DOC_TYPE_LABELS = {
    "PHOTO":       "Ảnh đại diện",
    "ID_CARD":     "CMND / CCCD",
    "DEGREE":      "Bằng cấp",
    "CERTIFICATE": "Chứng chỉ",
    "CONTRACT":    "Hợp đồng",
    "PROFILE":     "Hồ sơ nhân viên",
    "OTHER":       "Khác",
}

router = APIRouter()

# ── Cấu hình cột Excel ───────────────────────────────────────
EXPORT_COLUMNS = [
    ("Mã nhân viên",        "employee_code"),
    ("Họ và tên",           "full_name"),
    ("Giới tính",           "gender"),
    ("Ngày sinh",           "birthday"),
    ("Số CMND/CCCD",        "private_code"),
    ("Ngày cấp",            "private_code_date"),
    ("Nơi cấp",             "private_code_place"),
    ("Quốc tịch",           "nationality"),
    ("Dân tộc",             "ethnicity"),
    ("Email",               "email"),
    ("Điện thoại",          "phone"),
    ("Di động",             "mobile"),
    ("Địa chỉ thường trú",  "home_address"),
    ("Địa chỉ hiện tại",    "current_address"),
    ("Phòng ban (tên)",         "_department_name"),
    ("Vị trí công việc (tên)", "_position_name"),
    ("Chức danh (tên)",        "_job_title_name"),
    ("Trạng thái LĐ",          "job_status"),
    ("Ngày vào làm",        "job_date_join"),
    ("Ngày thử việc",       "job_date_try"),
    ("Ngày chính thức",     "job_reldate_join"),
    ("Ngày nghỉ việc",      "job_date_out"),
    ("Lý do nghỉ",          "job_out_reason"),
    ("Hình thức lương",     "salary_method"),
    ("Mức lương (VNĐ)",     "salary_real"),
    ("Ghi chú",             "description"),
    ("Trạng thái TK",       "_is_active"),
]

IMPORT_REQUIRED = {"Mã nhân viên", "Họ và tên"}

import unicodedata

def _normalize(s: str) -> str:
    """Lowercase + remove diacritics for fuzzy matching"""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s.lower().strip())
    return "".join(c for c in nfkd if not unicodedata.combining(c))

# All keys already normalized (no accents, lowercase)
GENDER_MAP  = {"nam": "MALE", "nu": "FEMALE", "khac": "OTHER", "male": "MALE", "female": "FEMALE"}
STATUS_MAP  = {
    "thu viec": "PROBATION", "probation": "PROBATION",
    "chinh thuc": "OFFICIAL", "official": "OFFICIAL",
    "da nghi": "RESIGNED", "resigned": "RESIGNED",
    "cham dut": "TERMINATED", "terminated": "TERMINATED",
}
SALARY_MAP  = {
    "luong co dinh": "FIXED", "fixed": "FIXED",
    "theo cong": "TIMESHEET", "timesheet": "TIMESHEET",
    "khoan san pham": "PIECE", "piece": "PIECE",
}

def _date_str(val):
    if val is None:
        return ""
    if isinstance(val, (date,)):
        return val.strftime("%d/%m/%Y")
    return str(val)

def _build_excel(rows, dept_map, pos_map=None, title_map=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách nhân sự"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for col_idx, (header, _) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Data
    for row_idx, p in enumerate(rows, 2):
        for col_idx, (_, field) in enumerate(EXPORT_COLUMNS, 1):
            if field == "_department_name":
                val = dept_map.get(str(p.department_id), "")
            elif field == "_position_name":
                val = (pos_map or {}).get(str(p.position_id), "")
            elif field == "_job_title_name":
                val = (title_map or {}).get(str(p.job_title_id), "")
            elif field == "_is_active":
                val = "Hoạt động" if p.is_active else "Vô hiệu"
            elif field in ("birthday", "private_code_date", "job_date_join",
                           "job_date_try", "job_reldate_join", "job_date_out"):
                val = _date_str(getattr(p, field, None))
            elif field == "gender":
                raw = getattr(p, field, None) or ""
                val = {"MALE": "Nam", "FEMALE": "Nữ", "OTHER": "Khác"}.get(raw, raw)
            elif field == "job_status":
                raw = getattr(p, field, None) or ""
                val = {"PROBATION": "Thử việc", "OFFICIAL": "Chính thức",
                       "RESIGNED": "Đã nghỉ", "TERMINATED": "Chấm dứt"}.get(raw, raw)
            elif field == "salary_method":
                raw = getattr(p, field, None) or ""
                val = {"FIXED": "Lương cố định", "TIMESHEET": "Theo công",
                       "PIECE": "Khoán sản phẩm"}.get(raw, raw)
            else:
                val = getattr(p, field, None)
                if val is not None:
                    val = str(val) if not isinstance(val, (int, float)) else val
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Column widths
    col_widths = [14,22,10,13,16,13,24,12,10,24,14,14,30,30,20,22,18,14,13,13,13,13,24,16,16,30,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    return wb


# ════════════════════════════════════════════════════════════
# PERSONNEL - Nhân viên
# ════════════════════════════════════════════════════════════

@router.get("", response_model=PaginatedPersonnel, summary="Danh sách nhân viên")
async def list_personnel(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None, description="Tìm theo tên, mã, email"),
    department_id: Optional[uuid.UUID] = Query(None),
    job_status: Optional[str] = Query(None, description="PROBATION/OFFICIAL/RESIGNED/TERMINATED"),
    is_active: Optional[bool] = Query(None),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = select(Personnel)

    if search:
        like = f"%{search}%"
        q = q.where(or_(
            Personnel.full_name.ilike(like),
            Personnel.employee_code.ilike(like),
            Personnel.email.ilike(like),
            Personnel.phone.ilike(like),
        ))
    if department_id:
        q = q.where(Personnel.department_id == department_id)
    if job_status:
        q = q.where(Personnel.job_status == job_status)
    if is_active is not None:
        q = q.where(Personnel.is_active == is_active)

    # Count
    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(count_q)).scalar_one()

    # Paginate
    q = q.order_by(Personnel.full_name).offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(q)).scalars().all()

    return PaginatedPersonnel(
        items=rows,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=math.ceil(total / page_size) if total else 0,
    )


@router.get("/all", response_model=list[PersonnelOut], summary="Tất cả nhân viên (cho dropdown)")
async def list_personnel_all(
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    result = await db.execute(
        select(Personnel)
        .where(Personnel.is_active == True)
        .order_by(Personnel.full_name)
    )
    return result.scalars().all()


@router.get("/export", summary="Xuất danh sách nhân sự ra Excel")
async def export_personnel(
    search: Optional[str] = Query(None),
    department_id: Optional[uuid.UUID] = Query(None),
    job_status: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = select(Personnel)
    if search:
        like = f"%{search}%"
        q = q.where(or_(
            Personnel.full_name.ilike(like),
            Personnel.employee_code.ilike(like),
            Personnel.email.ilike(like),
        ))
    if department_id:
        q = q.where(Personnel.department_id == department_id)
    if job_status:
        q = q.where(Personnel.job_status == job_status)
    if is_active is not None:
        q = q.where(Personnel.is_active == is_active)

    rows = (await db.execute(q.order_by(Personnel.full_name))).scalars().all()

    depts = (await db.execute(select(Department))).scalars().all()
    dept_map = {str(d.id): d.name for d in depts}

    positions = (await db.execute(select(Position))).scalars().all()
    pos_map = {str(p.id): p.title for p in positions}

    job_titles = (await db.execute(select(PositionTitle))).scalars().all()
    title_map = {str(t.id): t.title for t in job_titles}

    wb = _build_excel(rows, dept_map, pos_map, title_map)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    filename = f"nhan_su_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/import-template", summary="Tải file mẫu nhập nhân sự")
async def download_import_template(current_user=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mẫu nhập nhân sự"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [h for h, _ in EXPORT_COLUMNS]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Dòng gợi ý
    hints = [
        "NV001", "Nguyễn Văn A", "Nam / Nữ / Khác", "dd/mm/yyyy",
        "012345678901", "dd/mm/yyyy", "Công an TP.HCM", "Việt Nam", "Kinh",
        "nv@email.com", "024...", "09x...", "Địa chỉ thường trú", "Địa chỉ hiện tại",
        "Tên phòng ban", "Tên vị trí (VD: Kỹ sư phần mềm)", "Tên chức danh (VD: Chuyên viên)",
        "Thử việc / Chính thức / Đã nghỉ / Chấm dứt",
        "dd/mm/yyyy", "dd/mm/yyyy", "dd/mm/yyyy", "dd/mm/yyyy", "",
        "Lương cố định / Theo công / Khoán sản phẩm", "10000000", "", "Hoạt động / Vô hiệu",
    ]
    hint_font = Font(color="94A3B8", italic=True, size=10)
    for col_idx, hint in enumerate(hints, 1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.border = border

    col_widths = [14,22,10,13,16,13,24,12,10,24,14,14,30,30,20,24,20,18,13,13,13,13,24,20,16,30,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_nhap_nhan_su.xlsx"},
    )


@router.post("/import", summary="Nhập nhân sự từ file Excel")
async def import_personnel(
    file: UploadFile = File(...),
    update_existing: bool = Query(False, description="Cập nhật nếu mã NV đã tồn tại"),
    has_hint_row: bool = Query(False, description="File từ mẫu hệ thống (có dòng gợi ý ở dòng 2)"),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file Excel (.xlsx, .xls)")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File Excel không hợp lệ")

    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    # Map header → column index (0-based)
    col = {h: i for i, h in enumerate(headers)}
    missing = IMPORT_REQUIRED - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"File thiếu cột bắt buộc: {', '.join(missing)}")

    # Load lookup maps: name → id
    depts = (await db.execute(select(Department))).scalars().all()
    dept_name_map = {_normalize(d.name): d.id for d in depts}

    positions = (await db.execute(select(Position))).scalars().all()
    pos_name_map = {_normalize(p.title): p.id for p in positions}

    job_titles = (await db.execute(select(PositionTitle))).scalars().all()
    title_name_map = {_normalize(t.title): t.id for t in job_titles}

    def get(row, name):
        idx = col.get(name)
        if idx is None:
            return None
        v = ws.cell(row, idx + 1).value
        return str(v).strip() if v is not None else None

    def parse_date(s):
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                from datetime import datetime as dt
                return dt.strptime(s.strip(), fmt).date()
            except Exception:
                pass
        return None

    def parse_salary(s):
        if not s:
            return None
        try:
            cleaned = str(s).replace(",", "").replace(".", "").replace(" ", "").strip()
            val = float(cleaned)
            return val if val > 0 else None
        except Exception:
            return None

    def clean_phone(s, maxlen=20):
        """Lấy số đầu tiên nếu có nhiều số phân cách bởi , / space"""
        if not s:
            return None
        first = s.replace("/", ",").replace(";", ",").split(",")[0].strip()
        return first[:maxlen] if first else None

    def trunc(s, maxlen):
        """Cắt chuỗi về đúng độ dài tối đa"""
        if s is None:
            return None
        return str(s)[:maxlen]

    created, updated, errors = 0, 0, []

    start_row = 3 if has_hint_row else 2
    for row_idx in range(start_row, ws.max_row + 1):
        try:
            emp_code = get(row_idx, "Mã nhân viên")
            full_name = get(row_idx, "Họ và tên")
            if not emp_code and not full_name:
                continue  # dòng trống
            if not emp_code:
                errors.append({"row": row_idx, "error": f"Thiếu Mã nhân viên (tên: {full_name})"})
                continue
            if not full_name:
                errors.append({"row": row_idx, "error": f"Thiếu Họ và tên (mã: {emp_code})"})
                continue

            dept_name  = get(row_idx, "Phòng ban (tên)")
            dept_id    = dept_name_map.get(_normalize(dept_name)) if dept_name else None
            pos_name   = get(row_idx, "Vị trí công việc (tên)")
            pos_id     = pos_name_map.get(_normalize(pos_name)) if pos_name else None
            title_name = get(row_idx, "Chức danh (tên)")
            title_id   = title_name_map.get(_normalize(title_name)) if title_name else None

            gender_raw     = _normalize(get(row_idx, "Giới tính") or "")
            job_status_raw = _normalize(get(row_idx, "Trạng thái LĐ") or "")
            salary_raw     = _normalize(get(row_idx, "Hình thức lương") or "")
            is_active_raw  = _normalize(get(row_idx, "Trạng thái TK") or "hoat dong")

            data = {
                "employee_code":      trunc(emp_code, 50),
                "full_name":          trunc(full_name, 255),
                "gender":             GENDER_MAP.get(gender_raw),
                "birthday":           parse_date(get(row_idx, "Ngày sinh")),
                "private_code":       trunc(get(row_idx, "Số CMND/CCCD"), 20),
                "private_code_date":  parse_date(get(row_idx, "Ngày cấp")),
                "private_code_place": trunc(get(row_idx, "Nơi cấp"), 255),
                "nationality":        trunc(get(row_idx, "Quốc tịch"), 100),
                "ethnicity":          trunc(get(row_idx, "Dân tộc"), 100),
                "email":              trunc(get(row_idx, "Email"), 255),
                "phone":              clean_phone(get(row_idx, "Điện thoại")),
                "mobile":             clean_phone(get(row_idx, "Di động")),
                "home_address":       get(row_idx, "Địa chỉ thường trú"),
                "current_address":    get(row_idx, "Địa chỉ hiện tại"),
                "department_id":      dept_id,
                "position_id":        pos_id,
                "job_title_id":       title_id,
                "job_status":         STATUS_MAP.get(job_status_raw),
                "job_date_join":      parse_date(get(row_idx, "Ngày vào làm")),
                "job_date_try":       parse_date(get(row_idx, "Ngày thử việc")),
                "job_reldate_join":   parse_date(get(row_idx, "Ngày chính thức")),
                "job_date_out":       parse_date(get(row_idx, "Ngày nghỉ việc")),
                "job_out_reason":     get(row_idx, "Lý do nghỉ"),
                "salary_method":      SALARY_MAP.get(salary_raw),
                "salary_real":        parse_salary(get(row_idx, "Mức lương (VNĐ)")),
                "description":        get(row_idx, "Ghi chú"),
                "is_active":          is_active_raw != "vo hieu",
            }

            # Dùng savepoint để lỗi 1 dòng không làm hỏng transaction
            try:
                async with db.begin_nested():
                    existing = (await db.execute(
                        select(Personnel).where(Personnel.employee_code == emp_code)
                    )).scalar_one_or_none()

                    if existing:
                        if update_existing:
                            for k, v in data.items():
                                if k != "employee_code" and v is not None:
                                    setattr(existing, k, v)
                            existing.updated_by = current_user.id
                            updated += 1
                        else:
                            errors.append({"row": row_idx, "error": f"Mã NV '{emp_code}' đã tồn tại (bỏ qua)"})
                    else:
                        clean = {k: v for k, v in data.items() if v is not None}
                        p = Personnel(**clean, created_by=current_user.id)
                        db.add(p)
                        created += 1
            except Exception as row_exc:
                errors.append({"row": row_idx, "error": f"Dòng {emp_code}: {str(row_exc)[:120]}"})

        except Exception as exc:
            errors.append({"row": row_idx, "error": f"Lỗi đọc dòng: {str(exc)[:120]}"})
            continue

    try:
        await db.commit()
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi lưu dữ liệu: {str(exc)}")

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "total_rows": created + updated + len(errors),
        "headers_found": headers[:5],  # debug: 5 cột đầu tìm thấy
    }


@router.post("", response_model=PersonnelOut, status_code=status.HTTP_201_CREATED, summary="Thêm nhân viên")
async def create_personnel(
    body: PersonnelCreate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    # Kiểm tra mã nhân viên trùng
    exists = (await db.execute(
        select(Personnel).where(Personnel.employee_code == body.employee_code)
    )).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=400, detail=f"Mã nhân viên '{body.employee_code}' đã tồn tại")

    p = Personnel(**body.model_dump(), created_by=current_user.id)
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


@router.get("/{personnel_id}", response_model=PersonnelDetail, summary="Chi tiết nhân viên")
async def get_personnel(
    personnel_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(
        select(Personnel).where(Personnel.id == personnel_id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    # Load contracts
    contracts = (await db.execute(
        select(EmployeeContract)
        .options(selectinload(EmployeeContract.contract_type))
        .where(EmployeeContract.personnel_id == personnel_id)
        .order_by(EmployeeContract.date_start.desc())
    )).scalars().all()

    result = PersonnelDetail.model_validate(p)
    result.contracts = [EmployeeContractOut.model_validate(c) for c in contracts]
    return result


@router.put("/{personnel_id}", response_model=PersonnelOut, summary="Cập nhật nhân viên")
async def update_personnel(
    personnel_id: uuid.UUID,
    body: PersonnelUpdate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(
        select(Personnel).where(Personnel.id == personnel_id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    for k, v in body.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    p.updated_by = current_user.id
    await db.commit()
    await db.refresh(p)
    return p


@router.delete("/{personnel_id}", status_code=status.HTTP_204_NO_CONTENT, summary="Xóa / vô hiệu hóa nhân viên")
async def delete_personnel(
    personnel_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(
        select(Personnel).where(Personnel.id == personnel_id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    p.is_active = False
    p.updated_by = current_user.id
    await db.commit()


# ════════════════════════════════════════════════════════════
# EMPLOYEE CONTRACTS - Hợp đồng lao động
# ════════════════════════════════════════════════════════════

@router.get("/{personnel_id}/contracts", response_model=list[EmployeeContractOut], summary="HĐ của nhân viên")
async def list_contracts(
    personnel_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = (await db.execute(
        select(EmployeeContract)
        .options(selectinload(EmployeeContract.contract_type))
        .where(EmployeeContract.personnel_id == personnel_id)
        .order_by(EmployeeContract.date_start.desc())
    )).scalars().all()
    return rows


@router.post("/{personnel_id}/contracts", response_model=EmployeeContractOut, status_code=201, summary="Thêm hợp đồng")
async def create_contract(
    personnel_id: uuid.UUID,
    body: EmployeeContractCreate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    # Kiểm tra nhân viên tồn tại
    p = (await db.execute(select(Personnel).where(Personnel.id == personnel_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    # Kiểm tra mã hợp đồng
    dup = (await db.execute(
        select(EmployeeContract).where(EmployeeContract.contract_code == body.contract_code)
    )).scalar_one_or_none()
    if dup:
        raise HTTPException(status_code=400, detail=f"Mã hợp đồng '{body.contract_code}' đã tồn tại")

    data = body.model_dump()
    data["personnel_id"] = personnel_id
    data["created_by"] = current_user.id
    c = EmployeeContract(**data)
    db.add(c)
    await db.commit()
    await db.refresh(c)

    # Reload with relations
    c = (await db.execute(
        select(EmployeeContract)
        .options(selectinload(EmployeeContract.contract_type))
        .where(EmployeeContract.id == c.id)
    )).scalar_one()
    return c


@router.put("/contracts/{contract_id}", response_model=EmployeeContractOut, summary="Cập nhật hợp đồng")
async def update_contract(
    contract_id: uuid.UUID,
    body: EmployeeContractUpdate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    c = (await db.execute(
        select(EmployeeContract).where(EmployeeContract.id == contract_id)
    )).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng")

    for k, v in body.model_dump(exclude_none=True).items():
        setattr(c, k, v)
    await db.commit()
    await db.refresh(c)

    c = (await db.execute(
        select(EmployeeContract)
        .options(selectinload(EmployeeContract.contract_type))
        .where(EmployeeContract.id == contract_id)
    )).scalar_one()
    return c


@router.delete("/contracts/{contract_id}", status_code=204, summary="Xóa hợp đồng")
async def delete_contract(
    contract_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    c = (await db.execute(
        select(EmployeeContract).where(EmployeeContract.id == contract_id)
    )).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng")
    await db.delete(c)
    await db.commit()


# ════════════════════════════════════════════════════════════
# POSITIONS - Vị trí công việc
# ════════════════════════════════════════════════════════════

@router.get("/positions/list", response_model=list[PositionOut], summary="Danh sách vị trí")
async def list_positions(
    is_active: Optional[bool] = Query(True),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = select(Position).order_by(Position.title)
    if is_active is not None:
        q = q.where(Position.is_active == is_active)
    rows = (await db.execute(q)).scalars().all()
    return rows


@router.post("/positions", response_model=PositionOut, status_code=201, summary="Thêm vị trí")
async def create_position(
    body: PositionCreate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    dup = (await db.execute(select(Position).where(Position.code == body.code))).scalar_one_or_none()
    if dup:
        raise HTTPException(status_code=400, detail=f"Mã vị trí '{body.code}' đã tồn tại")
    p = Position(**body.model_dump())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


@router.put("/positions/{position_id}", response_model=PositionOut, summary="Cập nhật vị trí")
async def update_position(
    position_id: uuid.UUID,
    body: PositionUpdate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(select(Position).where(Position.id == position_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy vị trí")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    await db.commit()
    await db.refresh(p)
    return p


@router.delete("/positions/{position_id}", status_code=204, summary="Xóa vị trí")
async def delete_position(
    position_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(select(Position).where(Position.id == position_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy vị trí")
    p.is_active = False
    await db.commit()


# ════════════════════════════════════════════════════════════
# POSITION TITLES - Chức danh
# ════════════════════════════════════════════════════════════

@router.get("/job-titles/list", response_model=list[PositionTitleOut], summary="Danh sách chức danh")
async def list_job_titles(
    is_active: Optional[bool] = Query(True),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = select(PositionTitle).order_by(PositionTitle.priority, PositionTitle.title)
    if is_active is not None:
        q = q.where(PositionTitle.is_active == is_active)
    rows = (await db.execute(q)).scalars().all()
    return rows


@router.post("/job-titles", response_model=PositionTitleOut, status_code=201, summary="Thêm chức danh")
async def create_job_title(
    body: PositionTitleCreate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = PositionTitle(**body.model_dump())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return p


@router.put("/job-titles/{title_id}", response_model=PositionTitleOut, summary="Cập nhật chức danh")
async def update_job_title(
    title_id: uuid.UUID,
    body: PositionTitleUpdate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(select(PositionTitle).where(PositionTitle.id == title_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy chức danh")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    await db.commit()
    await db.refresh(p)
    return p


@router.delete("/job-titles/{title_id}", status_code=204, summary="Xóa chức danh")
async def delete_job_title(
    title_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(select(PositionTitle).where(PositionTitle.id == title_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy chức danh")
    p.is_active = False
    await db.commit()


# ════════════════════════════════════════════════════════════
# CONTRACT TYPES - Loại hợp đồng
# ════════════════════════════════════════════════════════════

@router.get("/contract-types/list", response_model=list[ContractTypeOut], summary="Danh sách loại HĐ")
async def list_contract_types(
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = (await db.execute(
        select(ContractType).where(ContractType.is_active == True).order_by(ContractType.title)
    )).scalars().all()
    return rows


@router.post("/contract-types", response_model=ContractTypeOut, status_code=201, summary="Thêm loại HĐ")
async def create_contract_type(
    body: ContractTypeCreate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    dup = (await db.execute(select(ContractType).where(ContractType.code == body.code))).scalar_one_or_none()
    if dup:
        raise HTTPException(status_code=400, detail=f"Mã loại HĐ '{body.code}' đã tồn tại")
    c = ContractType(**body.model_dump())
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return c


@router.put("/contract-types/{ct_id}", response_model=ContractTypeOut, summary="Cập nhật loại HĐ")
async def update_contract_type(
    ct_id: uuid.UUID,
    body: ContractTypeUpdate,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    c = (await db.execute(select(ContractType).where(ContractType.id == ct_id))).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Không tìm thấy loại hợp đồng")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(c, k, v)
    await db.commit()
    await db.refresh(c)
    return c


# ════════════════════════════════════════════════════════════
# PERSONNEL DOCUMENTS - Tài liệu hồ sơ nhân viên
# ════════════════════════════════════════════════════════════

@router.get("/{personnel_id}/documents", summary="Danh sách tài liệu của nhân viên")
async def list_documents(
    personnel_id: uuid.UUID,
    doc_type: Optional[str] = Query(None),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = select(PersonnelDocument).where(PersonnelDocument.personnel_id == personnel_id)
    if doc_type:
        q = q.where(PersonnelDocument.doc_type == doc_type)
    q = q.order_by(PersonnelDocument.uploaded_at.desc())
    rows = (await db.execute(q)).scalars().all()
    return [
        {
            "id": str(r.id),
            "personnel_id": str(r.personnel_id),
            "doc_type": r.doc_type,
            "doc_type_label": DOC_TYPE_LABELS.get(r.doc_type, r.doc_type),
            "file_name": r.file_name,
            "file_url": r.file_url,
            "file_type": r.file_type,
            "file_size_bytes": r.file_size_bytes,
            "notes": r.notes,
            "source": r.source,
            "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None,
        }
        for r in rows
    ]


@router.post("/{personnel_id}/documents", status_code=201, summary="Upload tài liệu nhân viên")
async def upload_document(
    personnel_id: uuid.UUID,
    file: UploadFile = File(...),
    doc_type: str = Query("OTHER"),
    notes: Optional[str] = Query(None),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    p = (await db.execute(select(Personnel).where(Personnel.id == personnel_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File quá lớn (tối đa 20MB)")

    save_dir = os.path.join(_PERSONNEL_UPLOAD_DIR, str(personnel_id))
    os.makedirs(save_dir, exist_ok=True)

    ext = os.path.splitext(file.filename or "")[1].lower() or ".bin"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(save_dir, unique_name)
    with open(filepath, "wb") as f:
        f.write(content)

    doc = PersonnelDocument(
        personnel_id=personnel_id,
        doc_type=doc_type,
        file_name=file.filename or unique_name,
        file_url=f"/uploads/personnel/{personnel_id}/{unique_name}",
        file_type=file.content_type,
        file_size_bytes=len(content),
        notes=notes,
        source="UPLOAD",
        uploaded_by=current_user.id,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return {
        "id": str(doc.id),
        "file_name": doc.file_name,
        "file_url": doc.file_url,
        "doc_type": doc.doc_type,
        "doc_type_label": DOC_TYPE_LABELS.get(doc_type, doc_type),
        "file_size_bytes": doc.file_size_bytes,
        "uploaded_at": doc.uploaded_at.isoformat(),
    }


@router.delete("/documents/{doc_id}", status_code=204, summary="Xóa tài liệu")
async def delete_document(
    doc_id: uuid.UUID,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    doc = (await db.execute(select(PersonnelDocument).where(PersonnelDocument.id == doc_id))).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài liệu")

    # Xóa file vật lý nếu là file upload (không xóa file cũ được link)
    if doc.source == "UPLOAD":
        phys = _PERSONNEL_UPLOAD_DIR + doc.file_url.replace("/uploads/personnel", "")
        if os.path.exists(phys):
            os.remove(phys)

    await db.delete(doc)
    await db.commit()


@router.get("/documents/types", summary="Danh sách loại tài liệu")
async def list_doc_types(current_user=Depends(get_current_user)):
    return [{"value": k, "label": v} for k, v in DOC_TYPE_LABELS.items()]
