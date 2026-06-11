"""
Timesheet / Bảng chấm công endpoints

Ô chấm công (hours JSONB) chấp nhận:
  - số giờ:  {"1": 8.0, "5": 4.5}
  - mã ký hiệu: {"17": "CP", "20": "CV"}  (CP, NP, CV, CCT, NT, NV, OL, N, CT...)

Cấu trúc dòng / nhân viên / tháng:
  - row_index 0: dòng đơn (nhân sự chuyên môn)
  - row_index 1: "Lái máy"  + row_index 2: "Việc khác" (nhân sự lái máy / công nhật)
    → dòng TỔNG được tính tự động = tổng giờ 2 dòng
"""
from __future__ import annotations
import calendar
import io
import json
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from app.schemas.timesheet import TimesheetRowIn, TimesheetRowOut
from app.core.auth import get_current_user

router = APIRouter()

# Mã ký hiệu chấm công (giống chú thích trên bảng Excel)
KNOWN_CODES = [
    ("CP",  "Nghỉ có phép"),
    ("NP",  "Nghỉ không phép"),
    ("CV",  "Chờ việc"),
    ("CCT", "Chuyển công trình"),
    ("NT",  "Nghỉ tết"),
    ("NV",  "Nghỉ việc"),
    ("OL",  "Nghỉ ốm"),
    ("N",   "Nghỉ"),
    ("CT",  "Công tác"),
]


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _safe_hours(h) -> dict:
    """Đảm bảo hours là dict hợp lệ."""
    if not h:
        return {}
    if isinstance(h, str):
        try:
            return json.loads(h)
        except Exception:
            return {}
    return dict(h)


def _cell_num(v) -> float:
    """Giá trị số của một ô (mã ký hiệu → 0)."""
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _is_code(v) -> bool:
    """Ô là mã ký hiệu (CP, CV...) thay vì số giờ."""
    return isinstance(v, str) and v.strip() != "" and _cell_num(v) == 0


def _norm_hours(h: dict) -> dict:
    """Chuẩn hoá hours trước khi lưu: số > 0 hoặc mã ký hiệu in hoa."""
    out = {}
    for k, v in (h or {}).items():
        if v is None:
            continue
        if isinstance(v, (int, float)):
            if float(v) > 0:
                out[str(k)] = float(v)
            continue
        s = str(v).strip()
        if not s:
            continue
        try:
            f = float(s.replace(",", "."))
            if f > 0:
                out[str(k)] = f
        except ValueError:
            out[str(k)] = s.upper()[:8]
    return out


def _person_summary(lines: list[dict], num_days: int) -> dict:
    """
    Tổng hợp tháng của 1 nhân viên từ các dòng nhập liệu.
    lines: danh sách hours-dict của các dòng nhập (1 dòng đơn, hoặc Lái máy + Việc khác)
    """
    leave_cp = ot_hours = 0.0
    ot_days = lt8 = eq8 = gt8 = 0
    for d in range(1, num_days + 1):
        key = str(d)
        total = 0.0
        codes = []
        for hrs in lines:
            v = hrs.get(key)
            if v in (None, ""):
                continue
            if _is_code(v):
                codes.append(str(v).strip().upper())
            else:
                total += _cell_num(v)
        if total > 0:
            if total < 8:
                lt8 += 1
            elif total == 8:
                eq8 += 1
            else:
                gt8 += 1
            ot = max(0.0, total - 8)
            ot_hours += ot
            if ot > 2:
                ot_days += 1
        elif any(c == "CP" for c in codes):
            leave_cp += 1
    return {
        "leave": leave_cp,
        "ot_hours": round(ot_hours, 1),
        "ot_days": ot_days,
        "lt8": lt8, "eq8": eq8, "gt8": gt8,
        "total_days": lt8 + eq8 + gt8,
    }


# ─────────────────────────────────────────────────────────────
# List rows for a month
# ─────────────────────────────────────────────────────────────

@router.get("", response_model=list[TimesheetRowOut])
async def list_rows(
    year:    int = Query(...),
    month:   int = Query(...),
    dept_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    dept_clause = "AND p.department_id = :dept_id::uuid" if dept_id else ""
    rows = await db.execute(text(f"""
        SELECT tr.*,
               p.full_name, p.employee_code,
               p.position AS position_text,
               p.mobile, p.phone,
               d.name     AS department_name
        FROM timesheet_rows tr
        JOIN personnel p ON p.id = tr.personnel_id
        LEFT JOIN departments d ON d.id = p.department_id
        WHERE tr.year = :year AND tr.month = :month {dept_clause}
        ORDER BY p.full_name, tr.row_index
    """), {"year": year, "month": month, **({"dept_id": dept_id} if dept_id else {})})
    result = []
    for r in rows.fetchall():
        row = dict(r._mapping)
        row["hours"] = _safe_hours(row.get("hours"))
        row["leave_days"] = float(row.get("leave_days") or 0)
        result.append(row)
    return result


# ─────────────────────────────────────────────────────────────
# Bulk save (upsert) rows
# ─────────────────────────────────────────────────────────────

@router.post("/rows/bulk")
async def bulk_save(
    rows: list[TimesheetRowIn],
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    saved_ids = []
    for row in rows:
        hours_json = json.dumps(_norm_hours(row.hours))

        if row.id:
            await db.execute(text("""
                UPDATE timesheet_rows SET
                    row_label        = :row_label,
                    work_description = :work_description,
                    hours            = CAST(:hours AS JSONB),
                    leave_days       = :leave_days,
                    notes            = :notes,
                    updated_at       = NOW()
                WHERE id = :id
            """), {
                "id": str(row.id), "row_label": row.row_label,
                "work_description": row.work_description,
                "hours": hours_json, "leave_days": row.leave_days,
                "notes": row.notes,
            })
            saved_ids.append(str(row.id))
        else:
            result = await db.execute(text("""
                INSERT INTO timesheet_rows
                    (id, personnel_id, year, month, row_index,
                     row_label, work_description, hours, leave_days, notes)
                VALUES
                    (gen_random_uuid(), :pid, :year, :month, :row_index,
                     :row_label, :work_description, CAST(:hours AS JSONB), :leave_days, :notes)
                ON CONFLICT (personnel_id, year, month, row_index) DO UPDATE SET
                    row_label        = EXCLUDED.row_label,
                    work_description = EXCLUDED.work_description,
                    hours            = EXCLUDED.hours,
                    leave_days       = EXCLUDED.leave_days,
                    notes            = EXCLUDED.notes,
                    updated_at       = NOW()
                RETURNING id
            """), {
                "pid": str(row.personnel_id), "year": row.year, "month": row.month,
                "row_index": row.row_index, "row_label": row.row_label,
                "work_description": row.work_description,
                "hours": hours_json, "leave_days": row.leave_days, "notes": row.notes,
            })
            saved_ids.append(str(result.scalar()))

    await db.commit()
    return {"ok": True, "ids": saved_ids}


# ─────────────────────────────────────────────────────────────
# Delete a row
# ─────────────────────────────────────────────────────────────

@router.delete("/rows/{row_id}", status_code=204)
async def delete_row(
    row_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(
        text("DELETE FROM timesheet_rows WHERE id = :id RETURNING id"),
        {"id": str(row_id)}
    )
    if not r.fetchone():
        raise HTTPException(status_code=404, detail="Không tìm thấy dòng chấm công")
    await db.commit()


# ─────────────────────────────────────────────────────────────
# Export Excel — theo mẫu bảng chấm công công trường
# ─────────────────────────────────────────────────────────────

@router.get("/export")
async def export_excel(
    year:    int = Query(...),
    month:   int = Query(...),
    dept_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    num_days = calendar.monthrange(year, month)[1]

    dept_clause = "AND p.department_id = :dept_id::uuid" if dept_id else ""
    rows_q = await db.execute(text(f"""
        SELECT tr.*, p.full_name, p.employee_code, p.position AS position_text,
               p.mobile, p.phone, d.name AS department_name
        FROM timesheet_rows tr
        JOIN personnel p ON p.id = tr.personnel_id
        LEFT JOIN departments d ON d.id = p.department_id
        WHERE tr.year = :year AND tr.month = :month {dept_clause}
        ORDER BY p.full_name, tr.row_index
    """), {"year": year, "month": month, **({"dept_id": dept_id} if dept_id else {})})
    db_rows = [dict(r._mapping) for r in rows_q.fetchall()]
    for r in db_rows:
        r["hours"] = _safe_hours(r.get("hours"))

    # Nhân sự chưa có dòng chấm công vẫn xuất hiện trên bảng
    personnel_q = await db.execute(text(f"""
        SELECT p.id, p.full_name, p.employee_code, p.position,
               p.mobile, p.phone, d.name AS department_name
        FROM personnel p
        LEFT JOIN departments d ON d.id = p.department_id
        WHERE p.is_active = TRUE {dept_clause}
        ORDER BY p.full_name
    """), {**({"dept_id": dept_id} if dept_id else {})})
    all_p = {str(r.id): dict(r._mapping) for r in personnel_q.fetchall()}

    # Gom dòng theo nhân viên
    by_pid: dict[str, dict[int, dict]] = {}
    pinfo: dict[str, dict] = {}
    for r in db_rows:
        pid = str(r["personnel_id"])
        by_pid.setdefault(pid, {})[int(r["row_index"])] = r
        pinfo[pid] = {
            "full_name": r["full_name"], "position": r.get("position_text"),
            "phone": r.get("mobile") or r.get("phone"),
            "department_name": r.get("department_name"),
        }
    for pid, p in all_p.items():
        if pid not in by_pid:
            by_pid[pid] = {}
            pinfo[pid] = {
                "full_name": p["full_name"], "position": p.get("position"),
                "phone": p.get("mobile") or p.get("phone"),
                "department_name": p.get("department_name"),
            }

    ordered_pids = sorted(by_pid.keys(), key=lambda pid: pinfo[pid]["full_name"] or "")

    # ── Build Excel ──────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Chấm công T{month:02d}-{year}"

    thin = Side(style="thin", color="B8BFCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1A2744")
    sum_fill    = PatternFill("solid", fgColor="E8ECF4")
    tong_fill   = PatternFill("solid", fgColor="F1F5F9")
    ot_fill     = PatternFill("solid", fgColor="DCFCE7")
    code_fill   = PatternFill("solid", fgColor="FEF3C7")

    hf = Font(bold=True, color="FFFFFF", size=9)

    def hc(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = hf
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
        cell.fill = header_fill
        return cell

    def dc(r, c, v, fill=None, bold=False, align="center", size=9, color=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, size=size, color=color or "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = bdr
        if fill:
            cell.fill = fill
        return cell

    N_FIXED = 6  # STT, Họ tên, Chức danh, Chuyên môn/Máy, SĐT, Loại dòng
    SUM_HEADERS = ["Nghỉ có phép", "Tăng ca (giờ)", "Ngày TC >2h", "<8h", "8h", ">8h", "Tổng ngày công"]
    total_cols = N_FIXED + num_days + len(SUM_HEADERS)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(row=1, column=1, value=f"BẢNG CHẤM CÔNG NHÂN SỰ THÁNG {month:02d}-{year}")
    title.font = Font(bold=True, size=14, color="1A2744")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header
    fixed_headers = ["STT", "Họ và tên", "Chức danh", "Chuyên môn / Máy", "SĐT", ""]
    for ci, h in enumerate(fixed_headers, start=1):
        hc(2, ci, h)
    for d in range(1, num_days + 1):
        hc(2, N_FIXED + d, f"{d:02d}")
    for ci, h in enumerate(SUM_HEADERS, start=N_FIXED + num_days + 1):
        hc(2, ci, h)
    ws.row_dimensions[2].height = 34

    # Column widths
    widths = [4.5, 20, 11, 16, 11, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(N_FIXED + d)].width = 4
    sum_widths = [7, 8, 7, 5.5, 5.5, 5.5, 8]
    for i, w in enumerate(sum_widths):
        ws.column_dimensions[get_column_letter(N_FIXED + num_days + 1 + i)].width = w

    ws.freeze_panes = ws.cell(row=3, column=N_FIXED + 1)

    def write_day_cells(excel_row, hours_or_total, bold=False, fill=None):
        """hours_or_total: dict {day: value} — value là số hoặc mã."""
        for d in range(1, num_days + 1):
            v = hours_or_total.get(str(d))
            if v in (None, "", 0):
                dc(excel_row, N_FIXED + d, "", fill=fill)
            elif _is_code(v):
                dc(excel_row, N_FIXED + d, str(v).upper(), fill=code_fill, size=8)
            else:
                n = _cell_num(v)
                dc(excel_row, N_FIXED + d, n if n % 1 else int(n),
                   fill=ot_fill if n > 8 else fill, bold=bold)

    def write_summary_cells(excel_row, summ):
        base = N_FIXED + num_days + 1
        vals = [summ["leave"], summ["ot_hours"], summ["ot_days"],
                summ["lt8"], summ["eq8"], summ["gt8"], summ["total_days"]]
        for i, v in enumerate(vals):
            dc(excel_row, base + i, v if v else "", fill=sum_fill, bold=True)

    excel_row = 3
    stt = 0
    for pid in ordered_pids:
        idx_rows = by_pid[pid]
        info = pinfo[pid]
        machine = 1 in idx_rows or 2 in idx_rows
        stt += 1

        if machine:
            r1 = idx_rows.get(1, {})
            r2 = idx_rows.get(2, {})
            h1 = _safe_hours(r1.get("hours"))
            h2 = _safe_hours(r2.get("hours"))
            summ = _person_summary([h1, h2], num_days)

            # Dòng TỔNG (tính tự động)
            totals = {}
            for d in range(1, num_days + 1):
                key = str(d)
                t = (0 if _is_code(h1.get(key)) else _cell_num(h1.get(key) or 0)) + \
                    (0 if _is_code(h2.get(key)) else _cell_num(h2.get(key) or 0))
                if t > 0:
                    totals[key] = t
                else:
                    code = next((v for v in (h1.get(key), h2.get(key)) if _is_code(v)), None)
                    if code:
                        totals[key] = code

            top = excel_row
            machine_label = r1.get("row_label") or "Lái máy"
            for rr, (label, fill) in enumerate([("TỔNG", tong_fill), ("Lái máy", None), ("Việc khác", None)]):
                dc(excel_row + rr, 6, label, fill=fill, bold=(rr == 0), size=8)
            write_day_cells(top,     totals, bold=True, fill=tong_fill)
            write_day_cells(top + 1, h1)
            write_day_cells(top + 2, h2)
            write_summary_cells(top, summ)
            base = N_FIXED + num_days + 1
            for rr in range(1, 3):
                for i in range(len(SUM_HEADERS)):
                    dc(top + rr, base + i, "", fill=sum_fill)

            # Cột thông tin gộp 3 dòng
            for col, val, align in [
                (1, stt, "center"), (2, info["full_name"], "left"),
                (3, info.get("position") or "", "center"),
                (4, machine_label, "center"),
                (5, info.get("phone") or "", "center"),
            ]:
                ws.merge_cells(start_row=top, start_column=col, end_row=top + 2, end_column=col)
                dc(top, col, val, align=align, bold=(col == 2))
                for rr in range(1, 3):
                    ws.cell(row=top + rr, column=col).border = bdr
            excel_row += 3
        else:
            r0 = idx_rows.get(0, {})
            h0 = _safe_hours(r0.get("hours"))
            summ = _person_summary([h0], num_days)
            dc(excel_row, 1, stt)
            dc(excel_row, 2, info["full_name"], align="left", bold=True)
            dc(excel_row, 3, info.get("position") or "")
            dc(excel_row, 4, "Chuyên môn", size=8)
            dc(excel_row, 5, info.get("phone") or "")
            dc(excel_row, 6, "", )
            write_day_cells(excel_row, h0)
            write_summary_cells(excel_row, summ)
            excel_row += 1

        ws.row_dimensions[excel_row - 1].height = 15

    # Chú thích mã ký hiệu
    excel_row += 1
    legend = "   ·   ".join(f"{c}: {label}" for c, label in KNOWN_CODES)
    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=total_cols)
    lc = ws.cell(row=excel_row, column=1, value=legend)
    lc.font = Font(italic=True, size=8, color="475569")
    lc.alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"cham_cong_T{month:02d}_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
