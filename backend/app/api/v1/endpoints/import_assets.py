"""
Import tài sản từ file Excel (.xlsx)
- Hỗ trợ file có ảnh nhúng trực tiếp trong ô
- Hỗ trợ file không có ảnh (chỉ có tên file)
- Tự động extract ảnh → lưu theo tên {asset_code}.jpg
"""
import io, uuid, json, re, unicodedata, zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.db.session import get_db
from app.core.auth import get_current_user

router = APIRouter()

import os
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "/app/uploads/assets"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ── Map trạng thái ────────────────────────────────────────────
STATUS_MAP = {
    'Đang sử dụng':  'IN_USE',
    'Chờ cấp phát':  'PENDING_ALLOCATION',
    'Hủy':           'CANCELLED',
    'Thanh lý':      'LIQUIDATED',
    'Đã thanh lý':   'LIQUIDATED',
    'Bảo trì':       'IN_MAINTENANCE',
    'Đang bảo trì':  'IN_MAINTENANCE',
    'Hỏng':          'BROKEN',
    'Mất':           'LOST',
    'Đã thu hồi':    'RECOVERED',
}

# ── Helpers ───────────────────────────────────────────────────
def _parse_date(val) -> Optional[date]:
    if not val: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    s = str(val).strip()
    if not s or s in ('None','HẾT HẠN','N/A','-','0',''): return None
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%m/%d/%Y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def _safe_int(val, default=0):
    try: return int(float(str(val))) if val not in (None,'','None') else default
    except: return default

def _safe_float(val, default=0.0):
    try: return float(str(val)) if val not in (None,'','None') else default
    except: return default

def _to_code(name):
    nfkd = unicodedata.normalize('NFKD', str(name))
    return re.sub(r'[^A-Z0-9]+', '_', nfkd.encode('ascii','ignore').decode().upper()).strip('_')[:50]

def _extract_year(s):
    m = re.search(r'\b(19|20)\d{2}\b', str(s))
    return m.group() if m else None


# ══════════════════════════════════════════════════════════════
#  CỐT LÕI: Giải thuật extract ảnh nhúng từ xlsx
#  Mapping: drawing anchor row → Mã TS → lưu {asset_code}.jpg
# ══════════════════════════════════════════════════════════════
def extract_embedded_images(xlsx_bytes: bytes) -> dict:
    """
    Trả về: { asset_code: (image_bytes, ext) }
    Giải thuật:
      1. Đọc drawing1.xml → map row_index → rId
      2. Đọc drawing rels → map rId → image filename
      3. Đọc sheet1.xml + sharedStrings → map row_index → Mã TS
      4. Ghép 3 map trên → { asset_code: image_bytes }
    """
    result = {}
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
            files = z.namelist()

            # Không có ảnh nhúng
            media_files = [f for f in files if f.startswith('xl/media/')]
            if not media_files:
                return {}

            # Bước 1: Gom tất cả rId → tên file ảnh từ tất cả drawing rels
            rid_to_file = {}
            for rels_path in files:
                if not (rels_path.startswith('xl/drawings/_rels/') and rels_path.endswith('.rels')):
                    continue
                with z.open(rels_path) as f:
                    root = ET.fromstring(f.read())
                for rel in root:
                    rid_to_file[rel.get('Id')] = rel.get('Target','').split('/')[-1]

            # Bước 2: row (0-based) → rId — quét tất cả drawing files
            row_to_rid = {}
            ns = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
            }
            for drawing_path in files:
                if not (drawing_path.startswith('xl/drawings/drawing') and drawing_path.endswith('.xml')):
                    continue
                with z.open(drawing_path) as f:
                    root = ET.fromstring(f.read())
                for anchor in (root.findall('xdr:oneCellAnchor', ns) +
                               root.findall('xdr:twoCellAnchor', ns)):
                    from_e = anchor.find('xdr:from', ns)
                    if from_e is None:
                        continue
                    row_el = from_e.find('xdr:row', ns)
                    if row_el is None:
                        continue
                    row = int(row_el.text)
                    blip = anchor.find('.//a:blip', ns)
                    if blip is not None and row not in row_to_rid:
                        rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        row_to_rid[row] = rid

            # Bước 3: row → Mã TS
            shared = []
            if 'xl/sharedStrings.xml' in files:
                with z.open('xl/sharedStrings.xml') as f:
                    ss = ET.fromstring(f.read())
                ns2 = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in ss.findall('s:si', ns2):
                    t = si.find('.//s:t', ns2)
                    shared.append(t.text if t is not None else '')

            with z.open('xl/worksheets/sheet1.xml') as f:
                sheet = ET.fromstring(f.read())
            ns3 = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            def _cell_value(cell):
                t = cell.get('t', '')
                if t == 'inlineStr':
                    is_e = cell.find('s:is', ns3)
                    t_e = is_e.find('s:t', ns3) if is_e is not None else None
                    return t_e.text if t_e is not None else None
                v = cell.find('s:v', ns3)
                if v is None:
                    return None
                if t == 's' and v.text:
                    return shared[int(v.text)]
                return v.text

            ma_ts_col = None
            rows_data = {}
            for row_elem in sheet.findall('.//s:row', ns3):
                ridx = int(row_elem.get('r')) - 1  # 0-based
                for cell in row_elem.findall('s:c', ns3):
                    ref = cell.get('r', '')
                    col = ''.join(c for c in ref if c.isalpha())
                    val = _cell_value(cell)
                    if val is None:
                        continue
                    if ridx == 0 and val == 'Mã TS':
                        ma_ts_col = col
                    if ridx > 0 and col == ma_ts_col and val:
                        rows_data[ridx] = str(val).strip()

            # Bước 4: Ghép → extract ảnh
            for row, rid in row_to_rid.items():
                asset_code = rows_data.get(row)
                if not asset_code:
                    continue
                img_fname = rid_to_file.get(rid)
                if not img_fname:
                    continue
                media_path = f'xl/media/{img_fname}'
                if media_path not in files:
                    continue
                img_bytes = z.read(media_path)
                ext = img_fname.rsplit('.',1)[-1].lower() if '.' in img_fname else 'jpg'
                result[asset_code] = (img_bytes, ext)

    except Exception as e:
        print(f"extract_embedded_images error: {e}")
    return result


# ── Reference table helpers ───────────────────────────────────
async def _load_cache(db, table, key_col, val_col='id'):
    res = await db.execute(text(f"SELECT {key_col}, {val_col} FROM {table}"))
    return {str(r[0]): str(r[1]) for r in res.fetchall()}

async def _get_or_create(db, table, code, name, extra, cache):
    if not code or not name: return None
    if code in cache: return cache[code]
    new_id = str(uuid.uuid4())
    cols = ['id','code','name'] + list(extra.keys())
    vals = [new_id, code, name] + list(extra.values())
    params = dict(zip(cols, vals))
    placeholders = ', '.join(f':{c}' for c in cols)
    await db.execute(
        text(f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({placeholders}) ON CONFLICT (code) DO NOTHING"),
        params
    )
    cache[code] = new_id
    return new_id


# ── Endpoints ─────────────────────────────────────────────────
@router.post("/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "Thiếu thư viện openpyxl.")

    content = await file.read()

    # Đếm ảnh nhúng
    images = extract_embedded_images(content)

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "File trống.")

    headers = [str(h).strip() if h else f'col_{i}' for i,h in enumerate(rows[0])]
    from collections import Counter
    status_count = Counter(str(dict(zip(headers,r)).get('Trạng thái','') or '') for r in rows[1:])

    samples = []
    for row in rows[1:6]:
        r = dict(zip(headers, row))
        samples.append({
            'ma_ts':     r.get('Mã TS'),
            'ten_ts':    r.get('Tên tài sản'),
            'trang_thai':r.get('Trạng thái'),
            'vi_tri':    r.get('Vị trí tài sản'),
            'nguyen_gia':r.get('Nguyên giá'),
            'has_image': r.get('Mã TS') in images,
        })

    # Analyse reference tables — which new records will be created
    loc_cache_c  = await _load_cache(db, 'locations',  'code')
    dept_cache_c = await _load_cache(db, 'departments','code')
    sup_cache_c  = await _load_cache(db, 'suppliers',  'code')
    at_cache_c   = await _load_cache(db, 'asset_types','code')
    existing_res = await db.execute(text("SELECT asset_code FROM assets"))
    existing_codes = {r[0] for r in existing_res.fetchall()}

    new_depts: dict = {}
    new_locs: dict = {}
    new_sups: dict = {}
    new_types: dict = {}
    to_import = to_skip = 0

    for row in rows[1:]:
        r = dict(zip(headers, row))
        code = str(r.get('Mã TS') or '').strip()
        name = str(r.get('Tên tài sản') or '').strip()
        if not code or not name:
            continue
        if code in existing_codes:
            to_skip += 1; continue
        to_import += 1

        for field, cache, store in [
            ('Phòng ban quản lý', dept_cache_c, new_depts),
            ('Vị trí tài sản',   loc_cache_c,  new_locs),
            ('Nhà cung cấp',     sup_cache_c,  new_sups),
        ]:
            val = str(r.get(field) or '').strip()
            if val and val != 'None':
                c = _to_code(val)
                if c not in cache and val not in store:
                    store[val] = c

        at_val = str(r.get('Loại TS') or r.get('Loại tài sản') or '').strip()
        if at_val and at_val != 'None':
            c = _to_code(at_val)
            if c not in at_cache_c and at_val not in new_types:
                new_types[at_val] = c

    return {
        'total_rows':      len(rows) - 1,
        'columns':         len(headers),
        'embedded_images': len(images),
        'status_counts':   dict(status_count),
        'samples':         samples,
        'to_import':       to_import,
        'to_skip':         to_skip,
        'new_departments': [{'name': k, 'code': v} for k, v in new_depts.items()],
        'new_locations':   [{'name': k, 'code': v} for k, v in new_locs.items()],
        'new_suppliers':   [{'name': k, 'code': v} for k, v in new_sups.items()],
        'new_asset_types': [{'name': k, 'code': v} for k, v in new_types.items()],
    }


@router.post("/import/execute")
async def execute_import(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "Thiếu thư viện openpyxl.")

    content = await file.read()

    # Extract ảnh nhúng trước
    embedded_images = extract_embedded_images(content)
    print(f"Extracted {len(embedded_images)} embedded images")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "File trống.")

    headers = [str(h).strip() if h else f'col_{i}' for i,h in enumerate(rows[0])]
    data_rows = [dict(zip(headers, row)) for row in rows[1:]]

    # Legal entity
    res = await db.execute(text("SELECT id FROM legal_entities WHERE code='DTH' LIMIT 1"))
    row = res.fetchone()
    le_id = str(row[0]) if row else str(uuid.uuid4())
    if not row:
        await db.execute(text("""
            INSERT INTO legal_entities (id, code, name, tax_code, address)
            VALUES (:id,'DTH','CÔNG TY TNHH ĐỒNG THUẬN HÀ','3702123456','Tây Ninh')
        """), {'id': le_id})

    # Caches
    loc_cache  = await _load_cache(db, 'locations',  'code')
    dept_cache = await _load_cache(db, 'departments','code')
    sup_cache  = await _load_cache(db, 'suppliers',  'code')
    at_cache   = await _load_cache(db, 'asset_types','code')
    grp_cache  = await _load_cache(db, 'asset_type_groups','code')

    async def get_location(name):
        if not name: return None
        name = str(name).strip()
        code = _to_code(name)
        if code in loc_cache: return loc_cache[code]
        n = name.upper()
        loc_type = 'OFFICE' if any(x in n for x in ['VĂN PHÒNG','TRƯỜNG CHINH']) else \
                   'WAREHOUSE' if 'KHO' in n else 'PROJECT_SITE'
        return await _get_or_create(db,'locations',code,name,
            {'legal_entity_id':le_id,'location_type':loc_type},loc_cache)

    async def get_dept(name):
        if not name: return None
        code = _to_code(str(name).strip())
        return await _get_or_create(db,'departments',code,str(name).strip(),
            {'legal_entity_id':le_id,'department_type':'ADMIN'},dept_cache)

    async def get_supplier(name, address=None, phone=None):
        if not name: return None
        code = _to_code(str(name).strip())
        extra = {}
        if address and str(address).strip() not in ('None', ''):
            extra['address'] = str(address).strip()[:500]
        if phone and str(phone).strip() not in ('None', ''):
            extra['phone'] = str(phone).strip()[:50]
        return await _get_or_create(db,'suppliers',code,str(name).strip(),extra,sup_cache)

    async def get_asset_type(loai_ts):
        if not loai_ts: return None
        loai_ts = str(loai_ts).strip()
        code = _to_code(loai_ts)
        if code in at_cache: return at_cache[code]
        n = loai_ts.upper()
        grp = ('CONSTRUCTION_MACHINE' if any(x in n for x in ['CƠ GIỚI','XE MÁY']) else
               'MEASUREMENT_DEVICE'   if 'ĐO ĐẠC' in n else
               'OFFICE_EQUIPMENT'     if 'VĂN PHÒNG' in n else
               'PPE'                  if 'BẢO HỘ' in n else 'TOOL')
        group_id = grp_cache.get(grp)
        if not group_id: return None
        return await _get_or_create(db,'asset_types',code,loai_ts,
            {'group_id':group_id},at_cache)

    # Existing codes
    res = await db.execute(text("SELECT asset_code FROM assets"))
    existing = {r[0] for r in res.fetchall()}

    imported = skipped = images_saved = 0
    errors = []

    for idx, r in enumerate(data_rows, start=2):
        try:
            code = str(r.get('Mã TS') or '').strip()
            name = str(r.get('Tên tài sản') or '').strip()
            if not code or not name: skipped += 1; continue
            if code in existing: skipped += 1; continue

            loc_id  = await get_location(r.get('Vị trí tài sản'))
            dept_id = await get_dept(r.get('Phòng ban quản lý'))
            sup_id  = await get_supplier(
                r.get('Nhà cung cấp'),
                address=r.get('Địa chỉ'),
                phone=r.get('Điện thoại'),
            )
            at_id   = await get_asset_type(r.get('Loại TS') or r.get('Loại tài sản'))

            status = STATUS_MAP.get(str(r.get('Trạng thái') or '').strip(), 'PENDING_ALLOCATION')

            nam_ns = str(r.get('Năm sản xuất - Nước sản xuất') or '').strip()
            year_mfg = _extract_year(nam_ns) if nam_ns and nam_ns != 'None' else None
            parts = re.split(r'[-–]', nam_ns) if nam_ns and nam_ns != 'None' else []
            country_mfg = parts[-1].strip()[:100] if len(parts) > 1 else (nam_ns[:100] if nam_ns and not re.match(r'^\d{4}$', nam_ns) else None)

            # Xử lý ảnh nhúng
            image_url = None
            if code in embedded_images:
                img_bytes, ext = embedded_images[code]
                img_filename = f"{code}.{ext}"
                img_path = UPLOAD_DIR / img_filename
                img_path.write_bytes(img_bytes)
                image_url = f"/api/v1/data/images/{img_filename}"
                images_saved += 1

            # Dynamic attrs
            dyn = {}
            anh_raw = str(r.get('Ảnh tài sản') or '').strip()
            if anh_raw and anh_raw != 'None':
                dyn['image_filename'] = anh_raw
            if r.get('Thông số thiết bị') and str(r['Thông số thiết bị']).strip() not in ('None',''):
                dyn['thong_so'] = str(r['Thông số thiết bị']).strip()
            if _safe_float(r.get('Thuê bao')) > 0:
                dyn['thue_bao'] = _safe_float(r['Thuê bao'])

            # Tags
            tags_raw = str(r.get('Nhãn tài sản') or '').strip()
            tags_list = [t.strip() for t in re.split(r'[,;]+', tags_raw) if t.strip()] if tags_raw and tags_raw != 'None' else []
            tags_pg = ('{' + ','.join('"' + t.replace('"', '\\"') + '"' for t in tags_list) + '}') if tags_list else None

            new_id = str(uuid.uuid4())
            await db.execute(text("""
                INSERT INTO assets (
                    id, asset_code, barcode, name,
                    legal_entity_id, asset_type_id,
                    managing_department_id, current_location_id, supplier_id,
                    model_series, year_manufactured, country_manufactured,
                    original_value, purchase_price, purchase_date,
                    report_increase_date, warranty_end_date, warranty_months,
                    expiry_date, depreciation_months, loan_amount,
                    quantity, qty_allocated, qty_recovered,
                    qty_maintenance, qty_liquidated, qty_lost, qty_cancelled, qty_broken,
                    status,
                    chassis_number, engine_number, registration_expiry,
                    description, condition_description, asset_image_url,
                    tags, dynamic_attributes
                ) VALUES (
                    :id, :code, :barcode, :name,
                    :le_id, :at_id,
                    :dept_id, :loc_id, :sup_id,
                    :model, :year, :country,
                    :orig_val, :purch_price, :purch_date,
                    :report_date, :warranty_end, :warranty_months,
                    :expiry_date, :dep_months, :loan_amt,
                    :qty, :qty_alloc, :qty_recv,
                    :qty_maint, :qty_liq, :qty_lost, :qty_cancel, :qty_broken,
                    :status,
                    :chassis, :engine, :reg_expiry,
                    :description, :condition_desc, :image_url,
                    CAST(:tags AS text[]), CAST(:dyn_attrs AS jsonb)
                ) ON CONFLICT (asset_code) DO NOTHING
            """), {
                'id': new_id, 'code': code,
                'barcode': str(r.get('Barcode') or '')[:100] or None,
                'name': name,
                'le_id': le_id, 'at_id': at_id,
                'dept_id': dept_id, 'loc_id': loc_id, 'sup_id': sup_id,
                'model': str(r.get('Model / Series') or '')[:200] or None,
                'year': year_mfg, 'country': country_mfg,
                'orig_val': _safe_float(r.get('Nguyên giá')),
                'purch_price': _safe_float(r.get('Giá mua')),
                'purch_date': _parse_date(r.get('Ngày mua')),
                'report_date': _parse_date(r.get('Ngày báo tăng')),
                'warranty_end': _parse_date(r.get('Ngày kết thúc BH')),
                'warranty_months': _safe_int(r.get('Thời gian BH (tháng)')),
                'expiry_date': _parse_date(r.get('Ngày hết hạn')),
                'dep_months': _safe_int(r.get('Khấu hao (tháng)')),
                'loan_amt': _safe_float(r.get('Đã vay')),
                'qty': max(1, _safe_int(r.get('Số lượng'), 1)),
                'qty_alloc': _safe_int(r.get('SL đã cấp phát')),
                'qty_recv': _safe_int(r.get('SL thu hồi')),
                'qty_maint': _safe_int(r.get('SL Bảo trì, Sửa chữa')),
                'qty_liq': _safe_int(r.get('SL thanh lý')),
                'qty_lost': _safe_int(r.get('SL mất')),
                'qty_cancel': _safe_int(r.get('SL huỷ')),
                'qty_broken': _safe_int(r.get('SL hỏng')),
                'status': status,
                'chassis': str(r.get('Số Khung') or '')[:100] or None,
                'engine': str(r.get('Số động cơ') or '')[:100] or None,
                'reg_expiry': _parse_date(r.get('Hạn đăng kiểm')),
                'description': str(r.get('Mô tả') or '')[:500] or None,
                'condition_desc': str(r.get('Diễn giải tình trạng') or '')[:500] or None,
                'image_url': image_url,
                'tags': tags_pg,
                'dyn_attrs': json.dumps(dyn, ensure_ascii=False),
            })
            existing.add(code)
            imported += 1

        except Exception as e:
            errors.append(f"Dòng {idx} ({r.get('Mã TS','?')}): {str(e)[:120]}")
            if len(errors) >= 20: errors.append("..."); break

    await db.commit()

    return {
        'success': True,
        'imported': imported,
        'skipped': skipped,
        'images_saved': images_saved,
        'errors': errors,
        'message': f'✅ Import {imported} tài sản, {images_saved} ảnh, bỏ qua {skipped} trùng.',
    }


# ── Serve ảnh ─────────────────────────────────────────────────
@router.get("/images/{filename}")
async def serve_image(filename: str):
    if '/' in filename or '..' in filename:
        raise HTTPException(400, "Tên file không hợp lệ")
    img_path = UPLOAD_DIR / filename
    if not img_path.exists():
        raise HTTPException(404, "Ảnh không tồn tại")
    ext = filename.rsplit('.',1)[-1].lower()
    media_type = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','webp':'image/webp'}.get(ext,'image/jpeg')
    return FileResponse(img_path, media_type=media_type,
                        headers={"Cache-Control": "public, max-age=86400"})


@router.get("/images/stats")
async def image_stats(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    img_files = list(UPLOAD_DIR.glob('*.*'))
    img_files = [f for f in img_files if f.suffix.lower() in ('.jpg','.jpeg','.png','.webp')]
    total_size = sum(f.stat().st_size for f in img_files)
    res = await db.execute(text(
        "SELECT COUNT(*) FROM assets WHERE asset_image_url IS NOT NULL AND asset_image_url != ''"
    ))
    return {
        'total_images': len(img_files),
        'total_size_mb': round(total_size/1024/1024, 2),
        'assets_with_image': res.scalar(),
        'upload_dir': str(UPLOAD_DIR),
    }
