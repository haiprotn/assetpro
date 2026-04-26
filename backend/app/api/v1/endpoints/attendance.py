"""
Attendance / Chấm công — Projects, daily records, work log allocation, report, export
"""
from __future__ import annotations
import calendar
import io
from datetime import date
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from app.schemas.attendance import (
    ProjectIn, ProjectOut,
    AttendanceUpsert, AttendanceOut,
    WorkLogIn, WorkLogOut,
)
from app.core.auth import get_current_user

router = APIRouter()


def _fmt_time(t) -> Optional[str]:
    """Convert time object → 'HH:MM' string, or None."""
    return t.strftime("%H:%M") if t else None


# ─────────────────────────────────────────────────────────────────────────────
# Projects (Công trình)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/projects", response_model=list[ProjectOut])
async def list_projects(
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("SELECT * FROM projects ORDER BY name"))
    return [dict(row._mapping) for row in r.fetchall()]


@router.post("/projects", response_model=ProjectOut)
async def create_project(
    body: ProjectIn,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("""
        INSERT INTO projects (id, code, name, description, is_active)
        VALUES (gen_random_uuid(), :code, :name, :description, :is_active)
        RETURNING *
    """), body.model_dump())
    await db.commit()
    return dict(r.fetchone()._mapping)


@router.put("/projects/{project_id}", response_model=ProjectOut)
async def update_project(
    project_id: UUID,
    body: ProjectIn,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("""
        UPDATE projects SET code=:code, name=:name,
            description=:description, is_active=:is_active
        WHERE id=:id RETURNING *
    """), {**body.model_dump(), "id": str(project_id)})
    row = r.fetchone()
    if not row:
        raise HTTPException(404, "Không tìm thấy công trình")
    await db.commit()
    return dict(row._mapping)


@router.delete("/projects/{project_id}", status_code=204)
async def delete_project(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(
        text("DELETE FROM projects WHERE id=:id RETURNING id"),
        {"id": str(project_id)},
    )
    if not r.fetchone():
        raise HTTPException(404, "Không tìm thấy công trình")
    await db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Attendance records
# ─────────────────────────────────────────────────────────────────────────────

@router.get("", response_model=list[AttendanceOut])
async def list_attendance(
    year:         int           = Query(...),
    month:        int           = Query(...),
    personnel_id: Optional[str] = Query(None),
    dept_id:      Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    num_days = calendar.monthrange(year, month)[1]
    params: dict = {
        "date_from": date(year, month, 1),
        "date_to":   date(year, month, num_days),
    }
    extra = ""
    if personnel_id:
        extra += " AND a.personnel_id = :personnel_id::uuid"
        params["personnel_id"] = personnel_id
    if dept_id:
        extra += " AND p.department_id = :dept_id::uuid"
        params["dept_id"] = dept_id

    r = await db.execute(text(f"""
        SELECT a.*,
               p.full_name, p.employee_code,
               p.position AS position_text,
               d.name     AS department_name
        FROM attendances a
        JOIN personnel p ON p.id = a.personnel_id
        LEFT JOIN departments d ON d.id = p.department_id
        WHERE a.date >= :date_from AND a.date <= :date_to {extra}
        ORDER BY p.full_name, a.date
    """), params)

    out = []
    for row in r.fetchall():
        rec = dict(row._mapping)
        rec["check_in"]  = _fmt_time(rec.get("check_in"))
        rec["check_out"] = _fmt_time(rec.get("check_out"))
        rec["total_hours"]    = float(rec.get("total_hours")    or 0)
        rec["overtime_hours"] = float(rec.get("overtime_hours") or 0)
        out.append(rec)
    return out


@router.put("/{personnel_id}/{att_date}")
async def upsert_attendance(
    personnel_id: UUID,
    att_date:     date,
    body:         AttendanceUpsert,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("""
        INSERT INTO attendances
            (id, personnel_id, date, check_in, check_out,
             total_hours, late_minutes, early_leave_minutes, overtime_hours,
             status, notes)
        VALUES
            (gen_random_uuid(), :pid, :date,
             CAST(NULLIF(:check_in,  '') AS TIME),
             CAST(NULLIF(:check_out, '') AS TIME),
             :total_hours, :late_minutes, :early_leave_minutes, :overtime_hours,
             :status, :notes)
        ON CONFLICT (personnel_id, date) DO UPDATE SET
            check_in              = EXCLUDED.check_in,
            check_out             = EXCLUDED.check_out,
            total_hours           = EXCLUDED.total_hours,
            late_minutes          = EXCLUDED.late_minutes,
            early_leave_minutes   = EXCLUDED.early_leave_minutes,
            overtime_hours        = EXCLUDED.overtime_hours,
            status                = EXCLUDED.status,
            notes                 = EXCLUDED.notes,
            updated_at            = NOW()
        RETURNING id
    """), {
        "pid":                str(personnel_id),
        "date":               att_date,
        "check_in":           body.check_in  or "",
        "check_out":          body.check_out or "",
        "total_hours":        body.total_hours,
        "late_minutes":       body.late_minutes,
        "early_leave_minutes": body.early_leave_minutes,
        "overtime_hours":     body.overtime_hours,
        "status":             body.status,
        "notes":              body.notes,
    })
    await db.commit()
    return {"ok": True, "id": str(r.scalar())}


# ─────────────────────────────────────────────────────────────────────────────
# Work logs (phân bổ giờ theo công trình)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/work-logs", response_model=list[WorkLogOut])
async def list_work_logs(
    year:         int           = Query(...),
    month:        int           = Query(...),
    personnel_id: Optional[str] = Query(None),
    project_id:   Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    num_days = calendar.monthrange(year, month)[1]
    params: dict = {
        "date_from": date(year, month, 1),
        "date_to":   date(year, month, num_days),
    }
    extra = ""
    if personnel_id:
        extra += " AND wl.personnel_id = :personnel_id::uuid"
        params["personnel_id"] = personnel_id
    if project_id:
        extra += " AND wl.project_id = :project_id::uuid"
        params["project_id"] = project_id

    r = await db.execute(text(f"""
        SELECT wl.*,
               pr.name AS project_name,
               pr.code AS project_code
        FROM work_logs wl
        JOIN projects pr ON pr.id = wl.project_id
        WHERE wl.date >= :date_from AND wl.date <= :date_to {extra}
        ORDER BY wl.date, wl.personnel_id, pr.name
    """), params)

    out = []
    for row in r.fetchall():
        rec = dict(row._mapping)
        rec["hours"] = float(rec.get("hours") or 0)
        out.append(rec)
    return out


@router.post("/work-logs", response_model=WorkLogOut)
async def create_work_log(
    body: WorkLogIn,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("""
        INSERT INTO work_logs (id, personnel_id, date, project_id, task_type, hours, notes)
        VALUES (gen_random_uuid(), :pid, :date, :project_id, :task_type, :hours, :notes)
        RETURNING id
    """), {
        "pid":        str(body.personnel_id),
        "date":       body.date,
        "project_id": str(body.project_id),
        "task_type":  body.task_type,
        "hours":      body.hours,
        "notes":      body.notes,
    })
    new_id = r.scalar()
    await db.commit()

    r2 = await db.execute(text("""
        SELECT wl.*, pr.name AS project_name, pr.code AS project_code
        FROM work_logs wl
        JOIN projects pr ON pr.id = wl.project_id
        WHERE wl.id = :id
    """), {"id": str(new_id)})
    rec = dict(r2.fetchone()._mapping)
    rec["hours"] = float(rec["hours"])
    return rec


@router.put("/work-logs/{log_id}", response_model=WorkLogOut)
async def update_work_log(
    log_id: UUID,
    body:   WorkLogIn,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(text("""
        UPDATE work_logs
        SET project_id=:project_id, task_type=:task_type,
            hours=:hours, notes=:notes, updated_at=NOW()
        WHERE id=:id RETURNING id
    """), {
        "project_id": str(body.project_id),
        "task_type":  body.task_type,
        "hours":      body.hours,
        "notes":      body.notes,
        "id":         str(log_id),
    })
    if not r.fetchone():
        raise HTTPException(404, "Không tìm thấy")
    await db.commit()

    r2 = await db.execute(text("""
        SELECT wl.*, pr.name AS project_name, pr.code AS project_code
        FROM work_logs wl
        JOIN projects pr ON pr.id = wl.project_id
        WHERE wl.id = :id
    """), {"id": str(log_id)})
    rec = dict(r2.fetchone()._mapping)
    rec["hours"] = float(rec["hours"])
    return rec


@router.delete("/work-logs/{log_id}", status_code=204)
async def delete_work_log(
    log_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    r = await db.execute(
        text("DELETE FROM work_logs WHERE id=:id RETURNING id"),
        {"id": str(log_id)},
    )
    if not r.fetchone():
        raise HTTPException(404, "Không tìm thấy")
    await db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Report data (monthly matrix)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/report")
async def get_report(
    year:    int           = Query(...),
    month:   int           = Query(...),
    dept_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    num_days = calendar.monthrange(year, month)[1]
    date_from = date(year, month, 1)
    date_to   = date(year, month, num_days)
    dept_clause = "AND p.department_id = :dept_id::uuid" if dept_id else ""
    params_dept = {"dept_id": dept_id} if dept_id else {}

    # ── attendance ──────────────────────────────────────────────
    att_r = await db.execute(text(f"""
        SELECT a.personnel_id, a.date, a.total_hours, a.overtime_hours,
               a.late_minutes, a.early_leave_minutes, a.status,
               a.check_in, a.check_out
        FROM attendances a
        JOIN personnel p ON p.id = a.personnel_id
        WHERE a.date >= :df AND a.date <= :dt {dept_clause}
    """), {"df": date_from, "dt": date_to, **params_dept})

    att_map: dict = {}   # { pid: { day_int: {...} } }
    for row in att_r.fetchall():
        rec = dict(row._mapping)
        pid = str(rec["personnel_id"])
        day = rec["date"].day
        att_map.setdefault(pid, {})[day] = {
            "total_hours":         float(rec.get("total_hours")    or 0),
            "overtime_hours":      float(rec.get("overtime_hours") or 0),
            "late_minutes":        int(rec.get("late_minutes")     or 0),
            "early_leave_minutes": int(rec.get("early_leave_minutes") or 0),
            "status":              rec.get("status", "present"),
            "check_in":            _fmt_time(rec.get("check_in")),
            "check_out":           _fmt_time(rec.get("check_out")),
        }

    # ── personnel list ──────────────────────────────────────────
    p_r = await db.execute(text(f"""
        SELECT p.id, p.full_name, p.employee_code, p.position,
               d.name AS department_name
        FROM personnel p
        LEFT JOIN departments d ON d.id = p.department_id
        WHERE p.is_active = TRUE {dept_clause}
        ORDER BY p.full_name
    """), params_dept)
    all_p = [dict(r._mapping) for r in p_r.fetchall()]

    # ── work logs ───────────────────────────────────────────────
    wl_r = await db.execute(text("""
        SELECT wl.personnel_id, wl.project_id, wl.hours,
               pr.name AS project_name, pr.code AS project_code
        FROM work_logs wl
        JOIN projects pr ON pr.id = wl.project_id
        WHERE wl.date >= :df AND wl.date <= :dt
    """), {"df": date_from, "dt": date_to})
    wl_map: dict = {}   # { pid: { proj_id: {hours, name, code} } }
    for row in wl_r.fetchall():
        rec = dict(row._mapping)
        pid  = str(rec["personnel_id"])
        prjid = str(rec["project_id"])
        wl_map.setdefault(pid, {}).setdefault(prjid, {
            "hours": 0, "name": rec["project_name"], "code": rec["project_code"],
        })
        wl_map[pid][prjid]["hours"] += float(rec.get("hours") or 0)

    # ── projects list ───────────────────────────────────────────
    proj_r = await db.execute(text(
        "SELECT id, code, name FROM projects WHERE is_active=TRUE ORDER BY name"
    ))
    projects = [{"id": str(r.id), "code": r.code, "name": r.name}
                for r in proj_r.fetchall()]

    # ── build employee summaries ────────────────────────────────
    employees = []
    for p in all_p:
        pid       = str(p["id"])
        days_data = att_map.get(pid, {})
        work_days = sum(1 for d in days_data.values()
                        if d["status"] == "present" and d["total_hours"] > 0)
        leave_days = sum(1 for d in days_data.values() if d["status"] == "leave")
        total_h   = sum(d["total_hours"]    for d in days_data.values())
        ot_h      = sum(d["overtime_hours"] for d in days_data.values())

        employees.append({
            "personnel_id":    pid,
            "full_name":       p["full_name"],
            "employee_code":   p.get("employee_code"),
            "department_name": p.get("department_name"),
            "position":        p.get("position"),
            "days":            {str(d): days_data[d] for d in sorted(days_data)},
            "summary": {
                "work_days":      work_days,
                "total_hours":    round(total_h, 1),
                "overtime_hours": round(ot_h, 1),
                "leave_days":     leave_days,
            },
            "work_logs": wl_map.get(pid, {}),
        })

    return {
        "year": year, "month": month, "num_days": num_days,
        "employees": employees, "projects": projects,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/export")
async def export_excel(
    year:    int           = Query(...),
    month:   int           = Query(...),
    dept_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as ddate

    report = await get_report(year=year, month=month, dept_id=dept_id, db=db, _user=_user)
    num_days  = report["num_days"]
    employees = report["employees"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Chấm công T{month:02d}-{year}"

    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="1A2744")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_fill    = PatternFill("solid", fgColor="1A2744")
    sub_fill    = PatternFill("solid", fgColor="E8ECF4")
    alt_fill    = PatternFill("solid", fgColor="F8FAFC")
    absent_fill = PatternFill("solid", fgColor="FEE2E2")
    leave_fill  = PatternFill("solid", fgColor="FEF3C7")
    ot_fill     = PatternFill("solid", fgColor="DCFCE7")
    wknd_fill   = PatternFill("solid", fgColor="2D4A8A")

    hf = Font(bold=True, color="FFFFFF", size=9)
    sf = Font(bold=True, size=9)

    def hc(r, c, v, bg=hdr_fill):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = hf if bg == hdr_fill else Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
        cell.fill = bg
        return cell

    def dc(r, c, v, fill=None, bold=False, align="center"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, size=9)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = bdr
        if fill:
            cell.fill = fill
        return cell

    sum_headers = ["Ngày công", "Tổng giờ", "Tăng ca (h)", "Nghỉ phép"]
    total_cols  = 4 + num_days + len(sum_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"BẢNG CHẤM CÔNG – THÁNG {month:02d}/{year}")
    title_cell.font      = Font(bold=True, size=14, color="1A2744")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row 2
    for ci, h in enumerate(["STT", "Họ và tên", "Bộ phận", "Chức vụ"], start=1):
        hc(2, ci, h)
    for d in range(1, num_days + 1):
        dow = ddate(year, month, d).weekday()
        bg  = wknd_fill if dow >= 5 else hdr_fill
        cell = hc(2, 4 + d, str(d), bg=bg)
        cell.value = f"{d}\n{'CN' if dow==6 else ['T2','T3','T4','T5','T6','T7'][dow]}"
    sum_fill = PatternFill("solid", fgColor="2D3F6B")
    for ci, h in enumerate(sum_headers, start=5 + num_days):
        hc(2, ci, h, bg=sum_fill)
    ws.row_dimensions[2].height = 32

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(4 + d)].width = 4.2
    for i in range(len(sum_headers)):
        ws.column_dimensions[get_column_letter(5 + num_days + i)].width = 9
    ws.freeze_panes = "E3"

    for stt, emp in enumerate(employees, 1):
        dr  = stt + 2
        bg  = alt_fill if stt % 2 == 0 else None
        dc(dr, 1, stt, fill=bg)
        dc(dr, 2, emp["full_name"],            fill=bg, align="left", bold=True)
        dc(dr, 3, emp.get("department_name") or "", fill=bg, align="left")
        dc(dr, 4, emp.get("position")        or "", fill=bg, align="left")

        for d in range(1, num_days + 1):
            day_rec = emp["days"].get(str(d))
            if day_rec:
                status = day_rec.get("status", "present")
                h      = day_rec["total_hours"]
                if status == "absent":
                    dc(dr, 4 + d, "V",  fill=absent_fill)
                elif status == "leave":
                    dc(dr, 4 + d, "NP", fill=leave_fill)
                elif h > 8:
                    dc(dr, 4 + d, h,    fill=ot_fill)
                else:
                    dc(dr, 4 + d, h if h > 0 else "", fill=bg)
            else:
                dc(dr, 4 + d, "", fill=bg)

        s    = emp["summary"]
        base = 5 + num_days
        dc(dr, base,     s["work_days"]      or "", fill=sub_fill, bold=True)
        dc(dr, base + 1, s["total_hours"]    or "", fill=sub_fill, bold=True)
        dc(dr, base + 2, s["overtime_hours"] or "", fill=sub_fill, bold=True)
        dc(dr, base + 3, s["leave_days"]     or "", fill=sub_fill, bold=True)
        ws.row_dimensions[dr].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cham_cong_T{month:02d}_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
